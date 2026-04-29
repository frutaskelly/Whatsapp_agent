"""Procesador de pedidos Excel.

Toma un Excel del cliente EHMO con hoja "BD" y genera el archivo de salida
con todas las hojas para Frutas y Verduras (Lote 5):
  1. BD                          - copia original sin modificar
  2. cambio lote 1 a lote 5      - lista de productos movidos
  3. BD cambio lote 1 a lote 5   - BD con lotes corregidos
  4. relacion de unidades a surtir - hospitales con si/no
  5. lista de compras            - consolidado total
  6. pedidos para surtir         - agrupado por hospital
  7+. una hoja por hospital

Basado en el script de CONTEXTO_SISTEMA_PEDIDOS.md.
"""
import json
import logging
import re
import time
from datetime import datetime
from functools import lru_cache
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from . import config
from .event_log import log_event

log = logging.getLogger(__name__)

# ─── Reglas de negocio ────────────────────────────────────────────────────────
EXCLUIDOS_KW = [
    "pichucalco", "palenque", "tila", "reforma",
    "yajalón", "yajalon", "amatán", "amatan",
]

CAMBIO_KW = [
    "ajo en bulbo", "ajonjolí", "ajonjoli",
    "cacahuate tostado sin sal",
    "canela en raja",
    "chile seco ancho", "chile seco guajillo", "chile seco pasilla",
    "epazote", "flor de jamaica",
    "nuez sin cascara", "nuez sin cáscara",  # cambio confirmado por Cristian 2026-04-27
    "orégano en hoja", "oregano en hoja",
    "perejil",
    "te de limón zacate", "te de limon zacate", "té de limón zacate",
    "te de manzanilla", "té de manzanilla",
    "te de yerbabuena", "té de yerbabuena",
    # Cambio Lote 1 → 5 confirmados por Cristian (2026-04-27, día 28)
    "palanqueta de cacahuate",
    "mermelada de fresa",
    "polvo para hornear",
]

IGNORAR_KW = [
    "almendra tostada",
    "salchicha",  # Lote 4 embutidos — no es nuestro (2026-04-27)
    # nota: "palanqueta de cacahuate" se movió a CAMBIO_KW (2026-04-27)
]


# ─── Listas extensibles desde storage/keywords.json ──────────────────────────
# Permite agregar nuevos productos al cambio / ignorar / hospitales excluidos
# SIN modificar código. El archivo es opcional. Si existe, sus entradas se
# concatenan con las constantes hardcoded de arriba (no las reemplaza).
#
# Estructura esperada (todos los campos opcionales):
# {
#   "cambio_kw": ["nuevo producto fyv"],
#   "ignorar_kw": ["producto a ignorar"],
#   "excluidos_kw": ["nuevo hospital a excluir"]
# }
def _keywords_file() -> Path:
    return config.BASE_DIR / "storage" / "keywords.json"


@lru_cache(maxsize=1)
def _extra_keywords() -> dict:
    p = _keywords_file()
    if not p.exists():
        return {}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {}
        return data
    except Exception as e:
        log.warning(f"No se pudo leer {p.name}: {e} — usando solo defaults")
        return {}


def recargar_keywords() -> dict:
    """Limpia el cache de keywords.json. Útil para aplicar cambios sin reiniciar."""
    _extra_keywords.cache_clear()
    extra = _extra_keywords()
    return {
        "cambio_kw_extra": len(extra.get("cambio_kw", [])),
        "ignorar_kw_extra": len(extra.get("ignorar_kw", [])),
        "excluidos_kw_extra": len(extra.get("excluidos_kw", [])),
    }


def _all_cambio_kw() -> list[str]:
    extra = _extra_keywords().get("cambio_kw", []) or []
    return CAMBIO_KW + [str(k).lower() for k in extra]


def _all_ignorar_kw() -> list[str]:
    extra = _extra_keywords().get("ignorar_kw", []) or []
    return IGNORAR_KW + [str(k).lower() for k in extra]


def _all_excluidos_kw() -> list[str]:
    extra = _extra_keywords().get("excluidos_kw", []) or []
    return EXCLUIDOS_KW + [str(k).lower() for k in extra]

# Catálogo de hospitales conocidos que SÍ surtimos. Mapea nombre canónico → lista
# de fingerprints (substrings case-insensitive) que identifican el hospital aunque
# venga con variantes (H.B.C., abreviaciones, sin acentos, etc.).
HOSPITALES_CONOCIDOS_SI = {
    "Hospital Básico Comunitario 12 Camas Berriozabal": ["berriozabal"],
    "Hospital Básico Comunitario Chiapa de Corzo": ["chiapa de corzo"],
    "Hospital Básico Comunitario de Cintalapa de Figueroa": ["cintalapa"],
    "Hospital Básico Comunitario Las Margaritas": ["margaritas"],
    "Hospital Básico Comunitario Manuel Velasco Suarez Acala": ["manuel velasco", "velasco suarez", "velasco suárez"],
    "Hospital Básico Comunitario Ángel Albino Corzo": ["ángel albino", "angel albino"],
    "Hospital Básico Comunitario Dr. Rafael Alfaro Gonzalez Pijijiapan": ["pijijiapan", "rafael alfaro"],
    "Hospital Básico de Frontera Comalapa": ["frontera comalapa", "comalapa"],
    "Hospital Chiapas nos une Dr. Jesús Gilberto Gomez Maza": ["chiapas nos une", "gomez maza", "gómez maza"],
    "Hospital de la Mujer Comitán": ["mujer comitán", "mujer comitan"],
    "Hospital de la Mujer San Cristóbal de las Casas": ["mujer san cristóbal", "mujer san cristobal"],
    "Hospital de las Culturas San Cristóbal de las Casas": ["culturas san cristóbal", "culturas san cristobal", "las culturas"],
    "Hospital General Bicentenario Villaflores": ["bicentenario villaflores", "bicentenario", "villaflores"],
    "Hospital General de Huixtla": ["huixtla"],
    "Hospital General de Ocosingo": ["ocosingo"],
    "Hospital General Dr. Juan C. Corzo Tonalá": ["juan c. corzo", "juan c corzo", "tonalá", "tonala"],
    "Hospital General Juárez Arriaga": ["juárez arriaga", "juarez arriaga"],
    "Hospital General María Ignacia Gandulfo Comitán": ["maría ignacia", "maria ignacia", "ignacia gandulfo", "gandulfo"],
    "Hospital General Tapachula": ["general tapachula", "hospital general tapachula"],
    "Hospital Regional Dr. Rafael Pascasio Gamboa Tuxtla": ["pascasio gamboa", "rafael pascasio"],
    "Unidad de Atención a la Salud Mental San Agustín": ["salud mental san agustín", "salud mental san agustin", "salud mental"],
}

# ─── Cliente: COMEDORES (6) ───────────────────────────────────────────────────
# Cliente independiente de EHMO. Compran lo mismo (frutas y verduras Lote 5)
# pero el formato de pedido es distinto: vienen en libreta/foto/voz, no Excel.
# Productos típicos: papas, cebollas, tomates, cilantro, chile jalapeño, chayotes,
# zanahorias, melones, sandías, ajo entero, huevos. Cantidades variables día a día.
# Confirmados por Cristian 2026-04-27.
COMEDORES_SI = {
    "Comedor Patria": ["patria", "comedor patria"],
    "Comedor CCI": ["cci", "comedor cci"],
    "Comedor 6 de Junio": ["6 de junio", "seis de junio", "comedor 6 de junio"],
    "Comedor Shanka": ["shanka", "comedor shanka"],
    "Comedor Jobo": ["jobo", "comedor jobo"],
    "Comedor Copoya": ["copoya", "comedor copoya"],
}

# Catálogo unificado de destinos: hospitales EHMO + comedores. Se usa para
# matching de nombres en pedidos, ajustes, modificaciones.
DESTINOS_CONOCIDOS_SI = {**HOSPITALES_CONOCIDOS_SI, **COMEDORES_SI}


def cliente_de(destino_canonico: str) -> str:
    """Devuelve el cliente al que pertenece un destino: 'EHMO' o 'COMEDORES'."""
    if destino_canonico in COMEDORES_SI:
        return "COMEDORES"
    return "EHMO"


NOMBRES_CORTOS = {
    "Hospital Básico Comunitario Ángel Albino Corzo": "Angel Albino Corzo",
    "Hospital Básico Comunitario Chiapa de Corzo": "Chiapa de Corzo",
    "Hospital Básico Comunitario de Cintalapa de Figueroa": "Cintalapa de Figueroa",
    "Hospital Básico Comunitario Las Margaritas": "Las Margaritas",
    "Hospital Básico Comunitario Manuel Velasco Suarez Acala": "Manuel Velasco Suarez Acala",
    "Hospital Chiapas nos une Dr. Jesús Gilberto Gomez Maza": "Jesús Gilberto Gomez Maza",
    "Hospital de la Mujer Comitán": "Mujer Comitán",
    "Hospital de la Mujer San Cristóbal de las Casas": "Mujer San Cristóbal de las Casa",
    "Hospital de las Culturas San Cristóbal de las Casas": "Culturas San Cristóbal de las C",
    "Hospital General María Ignacia Gandulfo Comitán": "María Ignacia Gandulfo ",
    "Hospital General Tapachula": " General Tapachula",
    "Hospital Regional Dr. Rafael Pascasio Gamboa Tuxtla": "Rafael Pascasio Gamboa Tuxtla",
    "Unidad de Atención a la Salud Mental San Agustín": "Salud Mental San Agustín",
    # Comedores
    "Comedor Patria": "Patria",
    "Comedor CCI": "CCI",
    "Comedor 6 de Junio": "6 de Junio",
    "Comedor Shanka": "Shanka",
    "Comedor Jobo": "Jobo",
    "Comedor Copoya": "Copoya",
}

MESES = {
    "enero": "ene", "febrero": "feb", "marzo": "mar", "abril": "abr",
    "mayo": "may", "junio": "jun", "julio": "jul", "agosto": "ago",
    "septiembre": "sep", "octubre": "oct", "noviembre": "nov", "diciembre": "dic",
}

# ─── Estilos ──────────────────────────────────────────────────────────────────
AZUL_OSC = "1F4E79"
AZUL_CLAR = "D6E4F0"
GRIS = "BBBBBB"


def _is_excluido(nombre):
    n = str(nombre).lower()
    return any(kw in n for kw in _all_excluidos_kw())


def _es_lote_5(lote_str) -> bool:
    """True si el lote es Lote 5 (Frutas y Verduras), aceptando con o sin
    prefijo numérico. EHMO a veces manda 'FRUTAS Y VERDURAS' sin el '5 ' inicial."""
    s = str(lote_str or "").strip().upper()
    return s in ("5 FRUTAS Y VERDURAS", "FRUTAS Y VERDURAS")


def _es_lote_1(lote_str) -> bool:
    """True si el lote es Lote 1 (Abarrotes), aceptando con o sin prefijo numérico.
    Excluye 'EXTRA ABARROTES' que es otra cosa (insumos, no comestibles)."""
    s = str(lote_str or "").strip().upper()
    return s in ("1 ABARROTES", "ABARROTES")


def _hospital_canonico(nombre: str) -> str | None:
    """Resuelve un nombre contra el catálogo de destinos (hospitales EHMO +
    comedores). Devuelve nombre canónico o None.

    Función nombrada "_hospital_canonico" por compatibilidad con código heredado;
    en realidad ahora resuelve cualquier destino conocido (no solo hospitales)."""
    n = nombre.lower()
    for canonical, fingerprints in DESTINOS_CONOCIDOS_SI.items():
        if any(fp in n for fp in fingerprints):
            return canonical
    return None


def _is_ignorar(alimento) -> bool:
    """True si el producto debe ser excluido del FyV aunque venga marcado como
    Lote 5 en el BD (ej. salchicha mal clasificada — pertenece a Lote 4 embutidos)."""
    a = str(alimento).lower()
    return any(kw in a for kw in _all_ignorar_kw())


def _is_cambio(alimento):
    a = str(alimento).lower()
    if any(kw in a for kw in _all_ignorar_kw()):
        return False
    return any(kw in a for kw in _all_cambio_kw())


def _get_nombre_corto(nombre):
    if nombre in NOMBRES_CORTOS:
        return NOMBRES_CORTOS[nombre]
    words = nombre.split()
    short = ""
    for w in reversed(words):
        candidate = (w + " " + short).strip() if short else w
        if len(candidate) <= 31:
            short = candidate
        else:
            break
    return short[:31]


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)


def _border():
    s = Side(style="thin", color=GRIS)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_row(ws, row_idx, values, bold=False, bg=None, font_color="000000"):
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", bold=bold, color=font_color, size=10)
        c.border = _border()
        c.alignment = Alignment(wrap_text=False)
        if bg:
            c.fill = _fill(bg)


def _extraer_fecha(filename: str, excel_path=None) -> str:
    """Extrae la fecha del archivo. Estrategia en orden de prioridad:

    1. Lee la celda A3 del Excel ("Suma de 28-abr" → "28 de abril").
       Es la fuente más confiable porque viene del cliente original.
    2. Match por nombre de archivo: 'Pedido 27 de abril original.xlsx' o
       'Pedido_general_28_de_abril.xlsx' (acepta espacios o underscores).
    3. Fallback: fecha de hoy.
    """
    # Estrategia 1: leer del contenido del Excel
    if excel_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                # Buscar "Suma de XX-mes" en las primeras 5 filas
                for r in range(1, 6):
                    for c in range(1, 5):
                        val = ws.cell(row=r, column=c).value
                        if not val or not isinstance(val, str):
                            continue
                        m = re.search(r"(\d{1,2})[\s\-/](\w{3,})", val.lower())
                        if m and "suma" in val.lower():
                            dia = m.group(1)
                            mes_kw = m.group(2)
                            # Mapear abr→abril, etc.
                            for nombre_largo in ["enero", "febrero", "marzo", "abril",
                                                  "mayo", "junio", "julio", "agosto",
                                                  "septiembre", "octubre", "noviembre",
                                                  "diciembre"]:
                                if nombre_largo.startswith(mes_kw[:3]):
                                    wb.close()
                                    return f"{dia} de {nombre_largo}"
            wb.close()
        except Exception:
            pass  # Caer a estrategia 2

    # Estrategia 2: nombre de archivo (espacios o underscores)
    # Normalizar separadores: '_' → ' '
    fname_norm = filename.replace("_", " ")
    m = re.search(r"Pedido\s+(.+?)(?:\s+original)?\s*\.xlsx", fname_norm, re.I)
    if m:
        # Limpiar palabras parásitas tipo "general", "genral" (typo)
        captured = m.group(1).strip()
        captured = re.sub(r"\b(general|genral)\b\s*", "", captured, flags=re.I).strip()
        return captured

    # Estrategia 3: hoy
    return datetime.now().strftime("%d de %B")


_MESES_NUMERO = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def fecha_a_iso(fecha_str: str | None, year: int | None = None) -> str | None:
    """Convierte '27 de abril' a '2026-04-27' (usa año actual si no se da uno).

    Devuelve None si no puede parsear. Útil para nombrar subcarpetas en Drive
    de forma sortable.
    """
    if not fecha_str:
        return None
    parts = fecha_str.lower().split()
    dia = None
    mes = None
    yr = year
    for p in parts:
        if dia is None:
            try:
                dia = int(p)
                continue
            except ValueError:
                pass
        if p in _MESES_NUMERO:
            mes = _MESES_NUMERO[p]
            continue
        if yr is None and p.isdigit() and len(p) == 4:
            yr = int(p)
    if dia is None or mes is None:
        return None
    yr = yr or datetime.now().year
    return f"{yr:04d}-{mes:02d}-{dia:02d}"


def procesar_pedido(input_excel: Path, output_dir: Path,
                    original_filename: str | None = None) -> dict | None:
    """Procesa un Excel EHMO con hoja BD.

    Devuelve un dict con:
      - output_path: Path del Excel generado
      - hospitales_si: list[str] hospitales que sí surtimos (incluidos en pedido)
      - hospitales_excluidos_detectados: list[str] de los 6 prohibidos que aparecieron
      - hospitales_desconocidos: list[str] que no están en regla 1 ni en regla 1b
      - productos_cambio_lote: list[str] productos movidos de Lote 1 a Lote 5
      - fecha: str de la fecha extraída

    Si el archivo no tiene hoja BD o el formato no coincide, devuelve None.
    `original_filename` se usa para extraer la fecha del nombre real del archivo
    (no del nombre local que puede tener prefijos de timestamp/phone).
    """
    input_excel = Path(input_excel)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log_event("processor", f"⚙️ Procesando {input_excel.name}")
    t0 = time.time()

    # ─── Leer y validar BD ───────────────────────────────────────────────────
    try:
        df_raw = pd.read_excel(input_excel, sheet_name="BD", header=0)
    except ValueError as e:
        log.warning(f"Excel no tiene hoja 'BD': {input_excel.name} ({e})")
        log_event("processor", f"⚠️ Sin hoja BD en {input_excel.name}", level="warn")
        return None
    except Exception as e:
        log.exception(f"Error leyendo Excel {input_excel.name}: {e}")
        log_event("processor", f"❌ Error leyendo Excel: {e}", level="error")
        return None

    if df_raw.shape[1] < 6:
        log.warning(f"Hoja BD con menos de 6 columnas: {df_raw.shape[1]}")
        return None

    if df_raw.shape[1] > 6:
        log.info(f"BD tiene {df_raw.shape[1]} columnas; uso las primeras 6 e ignoro el resto: {list(df_raw.columns[6:])}")
        df_raw = df_raw.iloc[:, :6]
    df_raw.columns = ["UNIDAD", "LOTE", "CBA", "ALIMENTO", "PRESENTACION", "CANTIDAD"]
    df_raw = df_raw.dropna(subset=["UNIDAD", "ALIMENTO"])
    df_raw["CANTIDAD"] = pd.to_numeric(df_raw["CANTIDAD"], errors="coerce").fillna(0)
    df_raw["UNIDAD"] = df_raw["UNIDAD"].astype(str).str.strip()
    df_raw["LOTE"] = df_raw["LOTE"].astype(str).str.strip()
    df_raw["ALIMENTO"] = df_raw["ALIMENTO"].astype(str).str.strip()

    # ─── Fecha ───────────────────────────────────────────────────────────────
    fecha_str = _extraer_fecha(original_filename or input_excel.name,
                                excel_path=input_excel)
    parts = fecha_str.lower().split()
    dia = parts[0] if parts else "fecha"
    mes_corto = MESES.get(parts[-1], parts[-1][:3]) if parts else "fec"
    fecha_corta = f"{dia}-{mes_corto}"

    # ─── Filtros ─────────────────────────────────────────────────────────────
    todos_hospitales = sorted(df_raw["UNIDAD"].unique())
    mask_excluido = df_raw["UNIDAD"].apply(_is_excluido)
    df_incluido = df_raw[~mask_excluido].copy()
    hospitales_si = sorted(df_incluido["UNIDAD"].unique())
    hospitales_excluidos_detectados = sorted(df_raw[mask_excluido]["UNIDAD"].unique())

    # Hospitales no listados en la regla 1b — el operador debe confirmar
    hospitales_desconocidos = sorted([h for h in hospitales_si if _hospital_canonico(h) is None])

    df_l5 = df_incluido[df_incluido["LOTE"].apply(_es_lote_5)].copy()

    # Filtrar productos mal clasificados como Lote 5 (ej. salchicha que debería
    # estar en Lote 4 EMBUTIDOS pero el BD del cliente a veces la pone en Lote 5).
    # IGNORAR_KW define la lista de productos que nunca debemos surtir como FyV.
    mal_clasificados = df_l5[df_l5["ALIMENTO"].apply(_is_ignorar)]
    if not mal_clasificados.empty:
        log_event("processor",
                  f"⚠️ {len(mal_clasificados)} producto(s) Lote 5 ignorado(s) — pertenecen a otro lote",
                  {"productos": mal_clasificados[["UNIDAD", "ALIMENTO", "CANTIDAD"]].to_dict("records")},
                  level="warn")
        df_l5 = df_l5[~df_l5["ALIMENTO"].apply(_is_ignorar)]

    # Detección de productos del cambio (CAMBIO_KW): se busca en CUALQUIER lote
    # que no sea Lote 5, incluyendo lote vacío, "EXTRA ABARROTES", o cualquier
    # categoría rara. EHMO a veces pone mermelada, perejil, etc. en lotes
    # equivocados o sin lote — los rescatamos siempre que coincidan con CAMBIO_KW.
    df_no_l5 = df_incluido[~df_incluido["LOTE"].apply(_es_lote_5)].copy()
    df_cambio = df_no_l5[df_no_l5["ALIMENTO"].apply(_is_cambio)].copy()
    productos_cambio_nombres = sorted(df_cambio["ALIMENTO"].unique())

    df_bd_corr = df_raw.copy()

    def corregir_lote(row):
        if (_is_cambio(row["ALIMENTO"])
                and not _es_lote_5(row["LOTE"])
                and not _is_excluido(row["UNIDAD"])):
            return "5 FRUTAS Y VERDURAS "
        return row["LOTE"]

    df_bd_corr["LOTE"] = df_bd_corr.apply(corregir_lote, axis=1)

    df_fyv = pd.concat([df_l5, df_cambio], ignore_index=True)
    # Filtros defensivos: solo cobramos lo que se surte. Descartamos:
    # - cantidades NaN o no-numéricas (ya convertidas a 0 por to_numeric arriba)
    # - cantidades 0 (líneas placeholder sin pedido real)
    # - cantidades negativas (cancelaciones)
    df_fyv = df_fyv[df_fyv["CANTIDAD"].notna() & (df_fyv["CANTIDAD"] > 0)]

    # Detectar y loguear cantidades sospechosas en el BD original (auditoría)
    sospechosos = df_raw[
        (df_raw["LOTE"].str.upper().str.contains("5 FRUTAS", na=False))
        & (df_raw["CANTIDAD"] <= 0)
    ]
    if not sospechosos.empty:
        log_event("processor",
                  f"⚠️ {len(sospechosos)} fila(s) Lote 5 con cantidad ≤ 0 (omitidas)",
                  {"ejemplos": sospechosos[["UNIDAD", "ALIMENTO", "CANTIDAD"]].head(3).to_dict("records")},
                  level="warn")

    # Hospitales surtibles que NO tienen FyV en este pedido (solo otros lotes)
    hospitales_con_fyv = sorted(df_fyv["UNIDAD"].unique())
    hospitales_si_sin_pedido_fyv = sorted([h for h in hospitales_si if h not in hospitales_con_fyv])

    # ─── Crear Excel de salida ───────────────────────────────────────────────
    output_path = output_dir / f"Pedido {fecha_str} Frutas y verduras.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja 1: BD original
    ws_bd = wb.create_sheet("BD")
    headers_bd = ["UNIDAD", "LOTE", "C.B.A", "ALIMENTO", "PRESENTACIÓN", fecha_str.upper()]
    _set_row(ws_bd, 1, headers_bd, bold=True, bg=AZUL_OSC, font_color="FFFFFF")
    for i, row in enumerate(df_raw.itertuples(index=False), 2):
        _set_row(ws_bd, i, [row.UNIDAD, row.LOTE, row.CBA, row.ALIMENTO, row.PRESENTACION, row.CANTIDAD])
    ws_bd.column_dimensions["A"].width = 60
    ws_bd.column_dimensions["D"].width = 60
    ws_bd.column_dimensions["F"].width = 18

    # Hoja 2: cambio lote
    ws_cambio = wb.create_sheet("cambio lote 1 a lote 5 ")
    _set_row(ws_cambio, 1, ["Productos cambio de Lote 1 a Lote 5"], bold=True, bg=AZUL_OSC, font_color="FFFFFF")
    for i, p in enumerate(productos_cambio_nombres, 2):
        c = ws_cambio.cell(row=i, column=1, value=p)
        c.font = _font()
        c.border = _border()
    ws_cambio.column_dimensions["A"].width = 60

    # Hoja 3: BD corregida
    ws_bdcorr = wb.create_sheet("BD cambio lote 1 a lote 5")
    _set_row(ws_bdcorr, 1, headers_bd, bold=True, bg=AZUL_OSC, font_color="FFFFFF")
    for i, row in enumerate(df_bd_corr.itertuples(index=False), 2):
        _set_row(ws_bdcorr, i, [row.UNIDAD, row.LOTE, row.CBA, row.ALIMENTO, row.PRESENTACION, row.CANTIDAD])
    ws_bdcorr.column_dimensions["A"].width = 60
    ws_bdcorr.column_dimensions["D"].width = 60
    ws_bdcorr.column_dimensions["F"].width = 18

    # Hoja 4: relación unidades — 3 estados:
    #   "si"          → tiene pedido FyV hoy, se surte
    #   "no"          → excluido permanente (Pichucalco, Palenque, etc.)
    #   "sin pedido"  → cliente válido pero hoy no pidió FyV (solo otros lotes)
    ws_rel = wb.create_sheet("relacion de unidades a surtir")
    _set_row(ws_rel, 1, ["Leyenda: si = surtir hoy · no = excluido permanente · sin pedido = cliente sin FyV hoy"],
             bold=True, bg=AZUL_OSC, font_color="FFFFFF")
    _set_row(ws_rel, 3, ["Etiquetas de fila", "unidades a surtir"], bold=True, bg=AZUL_CLAR)

    contadores = {"si": 0, "no": 0, "sin pedido": 0}
    for i, h in enumerate(sorted(todos_hospitales), 4):
        if h in hospitales_con_fyv:
            val = "si"
        elif _is_excluido(h):
            val = "no"
        else:
            val = "sin pedido"
        contadores[val] += 1
        _set_row(ws_rel, i, [h, val])

    last_row = len(todos_hospitales) + 4
    _set_row(ws_rel, last_row, ["Total general", len(todos_hospitales)], bold=True)
    _set_row(ws_rel, last_row + 1, [f"  · si (surtir hoy)", contadores["si"]])
    _set_row(ws_rel, last_row + 2, [f"  · no (excluido)", contadores["no"]])
    _set_row(ws_rel, last_row + 3, [f"  · sin pedido FyV hoy", contadores["sin pedido"]])
    ws_rel.column_dimensions["A"].width = 75
    ws_rel.column_dimensions["B"].width = 20

    # Hoja 4b: hospitales desconocidos (REVISAR — no están en el catálogo conocido)
    ws_desc = wb.create_sheet("hospitales desconocidos")
    if hospitales_desconocidos:
        _set_row(ws_desc, 1, ["⚠️ Estos hospitales NO están en tu catálogo. Confirma si los surtes."],
                 bold=True, bg=AZUL_OSC, font_color="FFFFFF")
        _set_row(ws_desc, 3, ["Hospital detectado", "Acción requerida"], bold=True, bg=AZUL_CLAR)
        for i, h in enumerate(hospitales_desconocidos, 4):
            _set_row(ws_desc, i, [h, "REVISAR"])
    else:
        _set_row(ws_desc, 1, ["✓ Todos los hospitales del pedido están en tu catálogo."],
                 bold=True, bg="00B050", font_color="FFFFFF")
    ws_desc.column_dimensions["A"].width = 70
    ws_desc.column_dimensions["B"].width = 20

    # Hoja 5: lista de compras
    ws_lista = wb.create_sheet("lista de compras")
    ws_lista.cell(row=1, column=1, value="LOTE").font = _font(bold=True)
    ws_lista.cell(row=1, column=2, value="5 FRUTAS Y VERDURAS ").font = _font()
    ws_lista.cell(row=2, column=1, value="UNIDAD").font = _font(bold=True)
    ws_lista.cell(row=2, column=2, value="(Varios elementos)").font = _font()
    _set_row(ws_lista, 4, ["Etiquetas de fila", f"Suma de {fecha_corta}"], bold=True, bg=AZUL_CLAR)
    compras = df_fyv.groupby("ALIMENTO")["CANTIDAD"].sum().reset_index()
    compras = compras[compras["CANTIDAD"] > 0].sort_values("ALIMENTO")
    for i, row in enumerate(compras.itertuples(index=False), 5):
        _set_row(ws_lista, i, [row.ALIMENTO, row.CANTIDAD])
    _set_row(ws_lista, len(compras) + 5, ["Total general", float(compras["CANTIDAD"].sum())], bold=True)
    ws_lista.column_dimensions["A"].width = 55
    ws_lista.column_dimensions["B"].width = 20

    # Hoja 6: pedidos para surtir
    ws_ped = wb.create_sheet("pedidos para surtir")
    ws_ped.cell(row=2, column=1, value="LOTE").font = _font(bold=True)
    ws_ped.cell(row=2, column=2, value="5 FRUTAS Y VERDURAS ").font = _font()
    _set_row(ws_ped, 4, ["Etiquetas de fila", f"Suma de {fecha_corta}"], bold=True, bg=AZUL_CLAR)
    hospitales_fyv = sorted(df_fyv["UNIDAD"].unique())
    row_idx = 5
    for hospital in hospitales_fyv:
        df_h = df_fyv[df_fyv["UNIDAD"] == hospital].groupby("ALIMENTO")["CANTIDAD"].sum().reset_index()
        df_h = df_h[df_h["CANTIDAD"] > 0].sort_values("ALIMENTO")
        _set_row(ws_ped, row_idx, [hospital, float(df_h["CANTIDAD"].sum())], bold=True, bg=AZUL_CLAR)
        row_idx += 1
        for _, prow in df_h.iterrows():
            _set_row(ws_ped, row_idx, [f"  {prow['ALIMENTO']}", prow["CANTIDAD"]])
            row_idx += 1
    _set_row(ws_ped, row_idx, ["Total general", float(df_fyv["CANTIDAD"].sum())], bold=True)
    ws_ped.column_dimensions["A"].width = 55
    ws_ped.column_dimensions["B"].width = 20

    # Hojas por hospital
    for hospital in hospitales_fyv:
        sname = _get_nombre_corto(hospital)
        ws_h = wb.create_sheet(sname)
        ws_h.cell(row=1, column=1, value="LOTE").font = _font(bold=True)
        ws_h.cell(row=1, column=2, value="5 FRUTAS Y VERDURAS ").font = _font()
        _set_row(ws_h, 3, ["Etiquetas de fila", f"Suma de {fecha_corta}"], bold=True, bg=AZUL_CLAR)
        df_h = df_fyv[df_fyv["UNIDAD"] == hospital].groupby("ALIMENTO")["CANTIDAD"].sum().reset_index()
        df_h = df_h[df_h["CANTIDAD"] > 0].sort_values("ALIMENTO")
        _set_row(ws_h, 4, [hospital, float(df_h["CANTIDAD"].sum())], bold=True, bg=AZUL_CLAR)
        for i, prow in enumerate(df_h.itertuples(index=False), 5):
            _set_row(ws_h, i, [prow.ALIMENTO, prow.CANTIDAD])
        ws_h.column_dimensions["A"].width = 55
        ws_h.column_dimensions["B"].width = 20

    wb.save(output_path)
    excel_elapsed = round((time.time() - t0) * 1000)
    log.info(f"Excel procesado guardado: {output_path}")

    # ─── PDF para imprimir (una hoja por hospital) ───────────────────────────
    from .pedido_pdf import generar_pdf_pedido
    pdf_path = output_dir / f"Pedido {fecha_str} Frutas y verduras.pdf"
    try:
        generar_pdf_pedido(df_fyv, fecha_str, pdf_path)
    except Exception as e:
        log.exception(f"Error generando PDF: {e}")
        log_event("processor", f"⚠️ PDF falló: {e}", level="warn")
        pdf_path = None

    # ─── Lista de compras consolidada (PDF imprimible + Excel editable) ──────
    from .lista_compras_pdf import generar_lista_compras_pdf, generar_lista_compras_xlsx
    lista_compras_path = output_dir / f"Lista de Compras {fecha_str}.pdf"
    lista_compras_xlsx_path = output_dir / f"Lista de Compras {fecha_str}.xlsx"
    try:
        generar_lista_compras_pdf(df_fyv, fecha_str, lista_compras_path)
    except Exception as e:
        log.exception(f"Error generando lista de compras PDF: {e}")
        log_event("processor", f"⚠️ Lista de compras PDF falló: {e}", level="warn")
        lista_compras_path = None
    try:
        generar_lista_compras_xlsx(df_fyv, fecha_str, lista_compras_xlsx_path)
    except Exception as e:
        log.exception(f"Error generando lista de compras Excel: {e}")
        log_event("processor", f"⚠️ Lista de compras Excel falló: {e}", level="warn")
        lista_compras_xlsx_path = None

    # ─── Notas de remisión con precios (una por hospital) ────────────────────
    from .nota_remision import generar_notas_remision
    notas_path = output_dir / f"Notas Remisión {fecha_str} Frutas y verduras.pdf"
    notas_info = None
    try:
        notas_info = generar_notas_remision(df_fyv, fecha_str, notas_path)
    except Exception as e:
        log.exception(f"Error generando notas de remisión: {e}")
        log_event("processor", f"⚠️ Notas remisión fallaron: {e}", level="warn")
        notas_path = None

    # ─── Persistir estado del día (para ajustes posteriores) ─────────────────
    from .estado_pedido import guardar_estado
    fecha_iso_calc = fecha_a_iso(fecha_str)
    if fecha_iso_calc:
        try:
            folios_iniciales = (notas_info or {}).get("folios", {})
            guardar_estado(fecha_iso_calc, fecha_str, df_fyv, folios=folios_iniciales)
        except Exception as e:
            log.exception(f"Error guardando estado del día: {e}")

    # ─── Relación de documentos del día (Excel + PDF horizontal) ────────────
    relacion_path = None
    relacion_pdf_path = None
    if fecha_iso_calc:
        try:
            from .relacion_documentos import generar_relacion_dia, generar_relacion_dia_pdf
            rel = generar_relacion_dia(fecha_iso_calc, fecha_legible=fecha_str)
            if rel and not rel.get("error"):
                relacion_path = rel["output_path"]
            rel_pdf = generar_relacion_dia_pdf(fecha_iso_calc, fecha_legible=fecha_str)
            if rel_pdf and not rel_pdf.get("error"):
                relacion_pdf_path = rel_pdf["output_path"]
        except Exception as e:
            log.exception(f"Error generando relación del día: {e}")
            log_event("processor", f"⚠️ Relación del día falló: {e}", level="warn")

    elapsed = round((time.time() - t0) * 1000)
    log_event("processor", f"✓ Excel generado ({len(wb.sheetnames)} hojas) en {excel_elapsed}ms",
              {"output": output_path.name, "sheets": len(wb.sheetnames),
               "hospitales_si": len(hospitales_si),
               "con_fyv": len(hospitales_con_fyv),
               "sin_fyv": len(hospitales_si_sin_pedido_fyv),
               "excluidos": len(hospitales_excluidos_detectados),
               "desconocidos": len(hospitales_desconocidos)})
    if hospitales_desconocidos:
        log.warning(f"Hospitales desconocidos en el pedido: {hospitales_desconocidos}")
        log_event("processor", f"⚠️ {len(hospitales_desconocidos)} hospital(es) desconocido(s)",
                  {"hospitales": hospitales_desconocidos}, level="warn")

    return {
        "output_path": output_path,
        "pdf_path": pdf_path,
        "lista_compras_path": lista_compras_path,
        "lista_compras_xlsx_path": lista_compras_xlsx_path,
        "notas_path": notas_path,
        "relacion_path": relacion_path,
        "relacion_pdf_path": relacion_pdf_path,
        "notas_total_general": notas_info["total_general"] if notas_info else None,
        "notas_sin_precio": notas_info["sin_precio_count"] if notas_info else 0,
        "hospitales_si": hospitales_si,
        "hospitales_con_fyv": hospitales_con_fyv,
        "hospitales_si_sin_pedido_fyv": hospitales_si_sin_pedido_fyv,
        "hospitales_excluidos_detectados": hospitales_excluidos_detectados,
        "hospitales_desconocidos": hospitales_desconocidos,
        "productos_cambio_lote": productos_cambio_nombres,
        "fecha": fecha_str,
    }


def extraer_fecha(filename: str) -> str:
    """Alias público para compatibilidad."""
    return _extraer_fecha(filename)
