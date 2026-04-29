"""Generador de Lista de Compras consolidada (PDF y Excel) para imprimir.

Diferente del PDF imprimible (una hoja por hospital) y de las notas de
remisión (con precios). Esta lista es UNA sola hoja con todos los
productos del día sumados — útil para que el equipo de compras vaya al
mayoreo con una sola referencia.

NO incluye precios ni desglose por hospital.
"""
import logging
from pathlib import Path
import pandas as pd

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.enums import TA_CENTER

from .display_names import corregir_nombre
from .event_log import log_event

log = logging.getLogger(__name__)

AZUL_OSC = colors.HexColor("#1F4E79")
AZUL_CLAR = colors.HexColor("#D6E4F0")
GRIS = colors.HexColor("#BBBBBB")
GRIS_CLARO = colors.HexColor("#F8F9FB")


def _fmt_qty(value) -> str:
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(f)) if f == int(f) else f"{f:.2f}"


def generar_lista_compras_pdf(df_fyv: pd.DataFrame, fecha_str: str,
                               output_path: Path) -> Path:
    """Genera PDF consolidado de productos a comprar al mayoreo.

    Suma cantidades por (alimento, presentación) sin desglosar por hospital.
    Sin precios — es para el equipo de compras.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df_fyv = df_fyv.copy()
    df_fyv["ALIMENTO"] = df_fyv["ALIMENTO"].apply(corregir_nombre)

    # Consolidar
    compras = (df_fyv
               .groupby(["ALIMENTO", "PRESENTACION"])["CANTIDAD"]
               .sum()
               .reset_index()
               .sort_values("ALIMENTO"))
    compras = compras[compras["CANTIDAD"] > 0]

    doc = SimpleDocTemplate(
        str(output_path), pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=f"Lista de Compras {fecha_str}",
    )

    styles = getSampleStyleSheet()
    style_h1 = ParagraphStyle(
        "h1", parent=styles["Heading1"], fontSize=16, leading=20,
        textColor=AZUL_OSC, alignment=TA_CENTER, spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"], fontSize=11, leading=13,
        alignment=TA_CENTER, textColor=colors.HexColor("#555555"),
        spaceAfter=12,
    )

    elements = [
        Paragraph(f"LISTA DE COMPRAS — {fecha_str.upper()}", style_h1),
        Paragraph(
            "Lote 5: Frutas y Verduras &nbsp;·&nbsp; "
            "Consolidado para compras al mayoreo",
            style_sub,
        ),
        Spacer(1, 0.3 * cm),
    ]

    # Tabla principal
    data = [["#", "ALIMENTO", "PRESENTACIÓN", "CANTIDAD"]]
    total = 0.0
    for i, row in enumerate(compras.itertuples(index=False), 1):
        cantidad = float(row.CANTIDAD)
        total += cantidad
        data.append([
            str(i),
            str(row.ALIMENTO),
            str(row.PRESENTACION),
            _fmt_qty(cantidad),
        ])

    # Fila final con total
    data.append([
        "",
        f"TOTAL DE PRODUCTOS DISTINTOS",
        f"{len(compras)} alimentos",
        _fmt_qty(total),
    ])

    table = Table(
        data,
        colWidths=[1 * cm, 11 * cm, 4 * cm, 2.5 * cm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), AZUL_OSC),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        # Body
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, GRIS_CLARO]),
        # Total
        ("BACKGROUND", (0, -1), (-1, -1), AZUL_CLAR),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
    ]))
    elements.append(table)

    doc.build(elements)
    log.info(f"Lista de compras PDF guardada: {output_path}")
    log_event("processor",
              f"📝 Lista de compras PDF generada ({len(compras)} productos)",
              {"output": output_path.name, "productos": len(compras)})
    return output_path


# ─── Versión Excel ────────────────────────────────────────────────────────────
_AZUL_OSC_HEX = "1F4E79"
_AZUL_CLAR_HEX = "D6E4F0"
_GRIS_HEX = "BBBBBB"


def generar_lista_compras_xlsx(df_fyv: pd.DataFrame, fecha_str: str,
                                output_path: Path) -> Path:
    """Genera Excel consolidado de productos a comprar al mayoreo.

    Misma información que el PDF pero en formato editable. Útil para que el
    equipo de compras pueda anotar cantidades reales compradas, proveedores, etc.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df_fyv = df_fyv.copy()
    df_fyv["ALIMENTO"] = df_fyv["ALIMENTO"].apply(corregir_nombre)

    compras = (df_fyv
               .groupby(["ALIMENTO", "PRESENTACION"])["CANTIDAD"]
               .sum()
               .reset_index()
               .sort_values("ALIMENTO"))
    compras = compras[compras["CANTIDAD"] > 0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"

    # Estilos
    azul_osc = PatternFill("solid", fgColor=_AZUL_OSC_HEX)
    azul_clar = PatternFill("solid", fgColor=_AZUL_CLAR_HEX)
    gris_claro = PatternFill("solid", fgColor="F8F9FB")
    border = Border(
        left=Side(style="thin", color=_GRIS_HEX),
        right=Side(style="thin", color=_GRIS_HEX),
        top=Side(style="thin", color=_GRIS_HEX),
        bottom=Side(style="thin", color=_GRIS_HEX),
    )

    # Título
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"LISTA DE COMPRAS — {fecha_str.upper()}")
    c.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1,
                value="Lote 5: Frutas y Verduras · Consolidado para compras al mayoreo")
    c.font = Font(name="Arial", italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["#", "ALIMENTO", "PRESENTACIÓN", "CANTIDAD",
               "PROVEEDOR", "OBSERVACIONES"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = azul_osc
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 28

    # Body
    total = 0.0
    for i, row in enumerate(compras.itertuples(index=False), 1):
        cantidad = float(row.CANTIDAD)
        total += cantidad
        excel_row = 4 + i
        cells = [
            (1, i, "center"),
            (2, str(row.ALIMENTO), "left"),
            (3, str(row.PRESENTACION), "center"),
            (4, cantidad if cantidad != int(cantidad) else int(cantidad), "right"),
            (5, "", "left"),  # Proveedor — vacío para llenar a mano
            (6, "", "left"),  # Observaciones — vacío
        ]
        bg = gris_claro if i % 2 == 0 else None
        for col, val, align in cells:
            c = ws.cell(row=excel_row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
            c.border = border
            if bg:
                c.fill = bg

    # Total row
    total_row = 4 + len(compras) + 1
    ws.cell(row=total_row, column=2, value="TOTAL DE PRODUCTOS DISTINTOS").font = Font(
        name="Arial", bold=True, size=11)
    ws.cell(row=total_row, column=3, value=f"{len(compras)} alimentos").font = Font(
        name="Arial", bold=True, size=11)
    val_total = total if total != int(total) else int(total)
    ws.cell(row=total_row, column=4, value=val_total).font = Font(
        name="Arial", bold=True, size=11)
    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = azul_clar
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal="center")
        elif col in (2,):
            cell.alignment = Alignment(horizontal="right")
        elif col in (3, 4):
            cell.alignment = Alignment(horizontal=("right" if col == 4 else "center"))

    # Anchos de columna
    widths = {1: 5, 2: 50, 3: 18, 4: 12, 5: 25, 6: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    log.info(f"Lista de compras Excel guardada: {output_path}")
    log_event("processor",
              f"📝 Lista de compras Excel generada ({len(compras)} productos)",
              {"output": output_path.name, "productos": len(compras)})
    return output_path
