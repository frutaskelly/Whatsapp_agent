"""Entry point del agente WhatsApp.

Local:        python app/main.py
Producción:   gunicorn --bind 0.0.0.0:$PORT app.main:app
"""
import logging
import sys
from pathlib import Path
from flask import Flask, jsonify, request, Response
from . import config
from .webhook import bp as webhook_bp
from . import message_log
from . import event_log
from .event_log import log_event

DASHBOARD_PATH = Path(__file__).parent / "dashboard.html"

# Forzar UTF-8 en stdout para que los emojis no revienten en consola Windows (cp1252)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)


def create_app():
    app = Flask(__name__)
    app.register_blueprint(webhook_bp)

    @app.route("/")
    def index():
        # Leer en cada request para reflejar ediciones sin reiniciar el server
        return Response(DASHBOARD_PATH.read_text(encoding="utf-8"), mimetype="text/html")

    @app.route("/api/status")
    def api_status():
        return jsonify({
            "service": "Frutas Kelly WhatsApp Agent",
            "status": "running",
            "environment": config.ENVIRONMENT,
        })

    @app.route("/api/messages")
    def api_messages():
        try:
            limit = int(request.args.get("limit", 200))
        except ValueError:
            limit = 200
        agent_id = (request.args.get("agent_id") or "").strip() or None
        msgs = message_log.read_messages(limit if not agent_id else 1000)
        if agent_id:
            msgs = [m for m in msgs if (m.get("meta") or {}).get("agent_id") == agent_id]
            msgs = msgs[-limit:]
        return jsonify({"messages": msgs, "filtered_agent_id": agent_id})

    @app.route("/api/events")
    def api_events():
        try:
            limit = int(request.args.get("limit", 200))
        except ValueError:
            limit = 200
        agent_id = (request.args.get("agent_id") or "").strip() or None
        evs = event_log.read_events(limit if not agent_id else 1000)
        if agent_id:
            evs = [e for e in evs if (e.get("meta") or {}).get("agent_id") == agent_id]
            evs = evs[-limit:]
        return jsonify({"events": evs, "filtered_agent_id": agent_id})

    @app.route("/api/consolidar-notas", methods=["POST", "GET"])
    def api_consolidar_notas():
        """Genera un PDF único con TODAS las notas vigentes del día."""
        from .processing_runner import _consolidar_notas
        return jsonify(_consolidar_notas("api") or {})

    @app.route("/api/reload-prices", methods=["POST", "GET"])
    def api_reload_prices():
        """Refresca la lista de precios desde el Excel (sin reiniciar Flask).
        Útil después de editar Lista_Precios_EHMO.xlsx para agregar productos."""
        from .pricing import cargar_lista_precios
        cargar_lista_precios.cache_clear()
        items = cargar_lista_precios()
        from . import config
        log_event("system", f"🔄 Lista de precios recargada: {len(items)} productos",
                  {"path": config.LISTA_PRECIOS_PATH})
        return jsonify({
            "ok": True,
            "productos_cargados": len(items),
            "ruta": config.LISTA_PRECIOS_PATH,
        })

    @app.route("/api/reload-keywords", methods=["POST", "GET"])
    def api_reload_keywords():
        """Refresca las listas de keywords (cambio_kw / ignorar_kw / excluidos_kw)
        desde storage/keywords.json. Útil después de agregar productos nuevos al
        archivo sin necesidad de reiniciar Flask."""
        from .pedido_processor import recargar_keywords
        stats = recargar_keywords()
        log_event("system", "🔄 Keywords extras recargadas", stats)
        return jsonify({"ok": True, **stats})

    @app.route("/api/relacion")
    def api_relacion():
        """Relación por DÍA (default) o semanal.

        Query params:
          - fecha: fecha-iso (YYYY-MM-DD). Si se da, genera relación del día.
          - semana + year: si no hay fecha, genera relación semanal completa.
        """
        from .relacion_documentos import generar_relacion_dia, generar_relacion_semanal
        from .drive_uploader import upload_file as drive_upload
        from datetime import datetime

        fecha = request.args.get("fecha")
        if fecha:
            result = generar_relacion_dia(fecha)
            if result.get("error"):
                return jsonify(result), 404
            drive_info = drive_upload(result["output_path"], subfolder=fecha)
            return jsonify({
                "ok": True,
                "tipo": "dia",
                "fecha": fecha,
                "hospitales": result["hospitales_count"],
                "total": result["total_general"],
                "output_local": str(result["output_path"]),
                "drive_link": (drive_info or {}).get("link"),
            })

        # Semanal (legacy / opcional)
        try:
            semana = int(request.args.get("semana") or datetime.now().isocalendar()[1] - 1)
            year = int(request.args.get("year") or datetime.now().year)
        except ValueError:
            return jsonify({"error": "semana/year inválidos"}), 400
        result = generar_relacion_semanal(semana, year)
        subfolder = f"SEM{semana} ({result['rango_inicio']} a {result['rango_fin']})"
        drive_info = drive_upload(result["output_path"], subfolder=subfolder)
        return jsonify({
            "ok": True,
            "tipo": "semana",
            "semana": semana,
            "rango": f"{result['rango_inicio']} a {result['rango_fin']}",
            "dias_con_data": result["dias_con_data"],
            "output_local": str(result["output_path"]),
            "drive_link": (drive_info or {}).get("link"),
        })

    @app.route("/api/simulate", methods=["POST"])
    def api_simulate():
        """Simula un mensaje entrante (texto y/o adjunto) y llama a Claude.
        No envía nada por WhatsApp, solo loguea ambas direcciones para el dashboard.
        Acepta JSON {message, phone} o multipart/form-data con 'message', 'phone' y 'file'.
        """
        from datetime import datetime
        from werkzeug.utils import secure_filename
        from .ai_agent import chat as ai_chat

        # Detectar formato del request
        ctype = request.content_type or ""
        uploaded_list = []
        agent_id = None
        if ctype.startswith("multipart/"):
            text = (request.form.get("message") or "").strip()
            phone = request.form.get("phone") or "simulator"
            uploaded_list = [f for f in request.files.getlist("file") if f and f.filename]
            agent_id = (request.form.get("agent_id") or "").strip() or None
        else:
            data = request.get_json(silent=True) or {}
            text = (data.get("message") or "").strip()
            phone = data.get("phone") or "simulator"
            agent_id = (data.get("agent_id") or "").strip() or None

        # Resolver agente activo (default si no se mandó uno)
        agente_activo = _resolver_agente(agent_id)

        log_event("webhook",
                  f"📩 Mensaje del simulador recibido"
                  + (f" ({len(uploaded_list)} archivos)" if len(uploaded_list) > 1 else ""),
                  {"phone": phone, "has_text": bool(text),
                   "n_files": len(uploaded_list),
                   "agent_id": agent_id,
                   "agente": agente_activo.get("nombre") if agente_activo else None})

        if not text and not uploaded_list:
            return jsonify({"error": "se requiere message o file"}), 400

        # ─── Procesar cada archivo (o solo texto si no hay archivos) ────────
        # Cada archivo se trata como un mensaje individual: se guarda, sube a
        # Drive, log "in", AI, log "out", maybe_process. Si vinieron varios
        # archivos juntos, todos comparten el mismo `text` del operador.
        from .drive_uploader import upload_file as drive_upload
        from .pedido_processor import _extraer_fecha, fecha_a_iso
        from .processing_runner import maybe_process

        def _handle_one(uploaded, file_idx: int, total: int) -> dict:
            attachment_path = None
            original_name = None
            drive_info = None
            if uploaded and uploaded.filename:
                original_name = uploaded.filename
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                safe = secure_filename(uploaded.filename) or "archivo"
                attachment_path = config.INBOX_DIR / f"{ts}_{phone}_{safe}"
                uploaded.save(attachment_path)
                log_event("storage",
                          f"💾 Archivo guardado: {original_name}"
                          + (f" ({file_idx}/{total})" if total > 1 else ""),
                          {"path": attachment_path.name,
                           "size_kb": attachment_path.stat().st_size // 1024,
                           "agent_id": agent_id})
                fecha_iso = fecha_a_iso(_extraer_fecha(original_name or ""))
                drive_info = drive_upload(attachment_path, original_name=original_name,
                                           subfolder=fecha_iso)

            # Log incoming (este archivo)
            in_body = text or ""
            in_meta = {"simulated": True}
            if attachment_path:
                lab = f"[adjunto: {attachment_path.name}]"
                if total > 1:
                    lab = f"[adjunto {file_idx}/{total}: {attachment_path.name}]"
                in_body = f"{in_body}\n{lab}".strip()
                in_meta["attachment"] = attachment_path.name
            if drive_info:
                in_meta["drive_link"] = drive_info["link"]
                in_meta["drive_id"] = drive_info["id"]
            if agente_activo:
                in_meta["agent_id"] = agente_activo.get("id")
                in_meta["agente"] = agente_activo.get("nombre")
            if total > 1:
                in_meta["batch_index"] = file_idx
                in_meta["batch_total"] = total
            message_log.log_message("in", phone, "text", in_body, in_meta)

            try:
                result = ai_chat(
                    phone, text, attachment_path=attachment_path,
                    agent_id=(agente_activo or {}).get("id"),
                    agent_addendum=(agente_activo or {}).get("system_prompt_addendum"),
                )
            except Exception as e:
                log.exception(f"Error en /api/simulate (archivo {file_idx}): {e}")
                err = f"Error: {e}"
                message_log.log_message("out", phone, "text", err,
                                          {"simulated": True, "error": True,
                                           "agent_id": (agente_activo or {}).get("id")})
                return {"error": str(e), "file": original_name}

            reply = result.get("respuesta_para_ehmo") or "(Claude no devolvió respuesta)"
            out_meta = {"simulated": True,
                        "intencion": result.get("intencion"),
                        "accion": result.get("accion")}
            if agente_activo:
                out_meta["agent_id"] = agente_activo.get("id")
                out_meta["agente"] = agente_activo.get("nombre")
            if total > 1:
                out_meta["batch_index"] = file_idx
                out_meta["batch_total"] = total
            message_log.log_message("out", phone, "text", reply, out_meta)

            processed = maybe_process(phone, attachment_path, result,
                                        original_filename=original_name)
            return {
                "file": original_name,
                "reply": reply,
                "intencion": result.get("intencion"),
                "accion": result.get("accion"),
                "datos": result.get("datos"),
                "processed": processed,
            }

        if not uploaded_list:
            # Solo texto, sin archivos — comportamiento original
            return jsonify(_handle_one(None, 1, 1))

        # Múltiples archivos: secuencial, devuelve array de resultados
        total = len(uploaded_list)
        if total > 1:
            log_event("webhook",
                      f"📦 Procesando {total} archivos en lote",
                      {"agent_id": agent_id, "phone": phone})
        results = []
        for i, up in enumerate(uploaded_list, 1):
            results.append(_handle_one(up, i, total))
        return jsonify({
            "batch": True,
            "total": total,
            "results": results,
        })

    @app.route("/health")
    def health():
        return jsonify({"status": "ok"}), 200

    # ─── Dashboard data endpoints ───────────────────────────────────────────
    @app.route("/api/dias-procesados")
    def api_dias_procesados():
        """Lista los días procesados con sus archivos generados.

        Para cada día (basado en storage/pedidos_dia/<fecha>.json) devuelve:
        fecha, fecha_legible, totales, # hospitales, # ajustes, archivos
        en storage/processed/ que matcheen con la fecha legible.
        """
        import json
        import re
        dias = []
        pedidos_dir = config.BASE_DIR / "storage" / "pedidos_dia"
        extras_dir = config.BASE_DIR / "storage" / "extras_dia"
        processed_dir = config.PROCESSED_DIR

        for state_path in sorted(pedidos_dir.glob("*.json"), reverse=True):
            try:
                state = json.loads(state_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            fecha_iso = state.get("fecha", state_path.stem)
            fecha_legible = state.get("fecha_legible", fecha_iso)

            total_reg = sum(h.get("total", 0) for h in state.get("hospitales", {}).values())
            n_hospitales = len(state.get("hospitales", {}))
            n_ajustes = len(state.get("ajustes", []))

            # Sumar extras del mismo día si existen
            ex_path = extras_dir / f"{fecha_iso}.json"
            total_ex = 0.0
            n_extras = 0
            if ex_path.exists():
                try:
                    ex = json.loads(ex_path.read_text(encoding="utf-8"))
                    total_ex = sum(e.get("importe", 0) for e in ex.get("extras", [])
                                    if e.get("cantidad", 0) > 0)
                    n_extras = sum(1 for e in ex.get("extras", []) if e.get("cantidad", 0) > 0)
                except Exception:
                    pass

            # Files: buscar archivos en processed/ cuyo nombre contenga la fecha legible
            # (ej. "27 de abril", "27 de April") o la fecha-iso
            archivos = []
            if processed_dir.exists():
                # Patrones posibles: "27 de abril", "27 de April", "(2026-04-27)"
                patterns = [fecha_legible, fecha_legible.replace("abril", "April"),
                            f"({fecha_iso})", fecha_iso]
                for f in sorted(processed_dir.iterdir()):
                    if not f.is_file():
                        continue
                    name = f.name
                    if any(p in name for p in patterns):
                        # Clasificar por tipo
                        nl = name.lower()
                        if "lista de compras" in nl or "lista_compras" in nl:
                            tipo = "lista_compras"
                        elif "nota" in nl and "remisión" in nl or "remision" in nl:
                            tipo = "nota_remision"
                        elif "relación" in nl or "relacion" in nl:
                            tipo = "relacion"
                        elif "extras" in nl:
                            tipo = "extras"
                        elif "pedido" in nl:
                            tipo = "pedido"
                        else:
                            tipo = "otro"
                        archivos.append({
                            "name": name,
                            "url": f"/files/processed/{name}",
                            "tipo": tipo,
                            "size_kb": round(f.stat().st_size / 1024, 1),
                            "modified": f.stat().st_mtime,
                        })

            # Drive folder URL — si tenemos GOOGLE_DRIVE_FOLDER_ID, link general
            drive_root = None
            if getattr(config, "GOOGLE_DRIVE_FOLDER_ID", None):
                drive_root = f"https://drive.google.com/drive/folders/{config.GOOGLE_DRIVE_FOLDER_ID}"

            dias.append({
                "fecha_iso": fecha_iso,
                "fecha_legible": fecha_legible,
                "hospitales": n_hospitales,
                "ajustes": n_ajustes,
                "extras": n_extras,
                "total_regular": round(total_reg, 2),
                "total_extras": round(total_ex, 2),
                "total_dia": round(total_reg + total_ex, 2),
                "archivos": archivos,
                "drive_folder": drive_root,
            })
        return jsonify({"dias": dias})

    @app.route("/api/relaciones-todas")
    def api_relaciones_todas():
        """Tabla consolidada de TODAS las remisiones (regulares + extras) en todos los días."""
        import json
        remisiones = []
        pedidos_dir = config.BASE_DIR / "storage" / "pedidos_dia"
        extras_dir = config.BASE_DIR / "storage" / "extras_dia"

        for state_path in sorted(pedidos_dir.glob("*.json"), reverse=True):
            try:
                state = json.loads(state_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            fecha_iso = state.get("fecha", state_path.stem)
            fecha_legible = state.get("fecha_legible", fecha_iso)
            for hospital, info in state.get("hospitales", {}).items():
                folio = info.get("folio_remision") or ""
                try:
                    folio_int = int(folio) if folio else 0
                except (TypeError, ValueError):
                    folio_int = 0
                remisiones.append({
                    "fecha_iso": fecha_iso,
                    "fecha_legible": fecha_legible,
                    "folio": folio_int,
                    "folio_str": folio,
                    "destino": hospital,
                    "tipo": "regular",
                    "estado": info.get("estado", "vigente"),
                    "total": round(info.get("total", 0), 2),
                    "productos": sum(1 for p in info.get("productos", [])
                                       if p.get("cantidad", 0) > 0),
                })

        # Extras al ALMACÉN EHMO u otros destinos
        for ex_path in sorted(extras_dir.glob("*.json"), reverse=True):
            try:
                ex_state = json.loads(ex_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            fecha_iso = ex_state.get("fecha", ex_path.stem)
            fecha_legible = ex_state.get("fecha_legible", fecha_iso)
            folios_destinos = ex_state.get("folios_por_destino") or {}
            estados_destinos = ex_state.get("estados_por_destino") or {}

            # Agrupar extras por destino
            por_destino: dict[str, list[dict]] = {}
            for e in ex_state.get("extras", []):
                if e.get("cantidad", 0) <= 0:
                    continue
                por_destino.setdefault(e.get("hospital", "ALMACÉN EHMO"), []).append(e)

            for destino, items in por_destino.items():
                folio = folios_destinos.get(destino) or ""
                try:
                    folio_int = int(folio) if folio else 0
                except (TypeError, ValueError):
                    folio_int = 0
                total = sum(e.get("importe", 0) for e in items)
                remisiones.append({
                    "fecha_iso": fecha_iso,
                    "fecha_legible": fecha_legible,
                    "folio": folio_int,
                    "folio_str": folio,
                    "destino": f"{destino} (EXTRA)",
                    "tipo": "extra",
                    "estado": estados_destinos.get(destino, "vigente"),
                    "total": round(total, 2),
                    "productos": len(items),
                })

        # Orden: fecha desc, luego folio asc
        remisiones.sort(key=lambda r: (r["fecha_iso"], -r["folio"] if r["folio"] else 0), reverse=True)
        return jsonify({"remisiones": remisiones})

    # ─── Listas de precios (catálogo multi-lista) ───────────────────────────
    def _resolver_lista_path(lista_id: str | None) -> "Path | None":
        """Devuelve el path xlsx de la lista pedida.

        Si lista_id es None, devuelve la lista por default (LISTA_PRECIOS_PATH).
        Resuelve desde agentes.json → listas_precios_archivos.
        """
        from pathlib import Path
        if not lista_id or lista_id == "default":
            return Path(config.LISTA_PRECIOS_PATH)
        data = _cargar_agentes_data()
        archivos = data.get("listas_precios_archivos") or {}
        rel = archivos.get(lista_id)
        if not rel:
            return None
        p = Path(rel)
        if not p.is_absolute():
            p = Path(config.BASE_DIR) / rel
        return p

    @app.route("/api/listas-precios", methods=["GET"])
    def api_listas_precios_catalog():
        """Lista las listas de precios disponibles (catálogo)."""
        import openpyxl
        from pathlib import Path
        data = _cargar_agentes_data()
        archivos = data.get("listas_precios_archivos") or {}
        meta = data.get("listas_precios_meta") or {}
        out = []
        for lid, rel in archivos.items():
            p = Path(rel)
            if not p.is_absolute():
                p = Path(config.BASE_DIR) / rel
            existe = p.exists()
            n_items = 0
            if existe:
                try:
                    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
                    ws = wb["Lista de Precios"] if "Lista de Precios" in wb.sheetnames else wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(v is not None for v in row): n_items += 1
                    wb.close()
                except Exception:
                    pass
            m = meta.get(lid, {})
            out.append({
                "id": lid,
                "archivo": rel,
                "existe": existe,
                "items": n_items,
                "nombre": m.get("nombre", lid),
                "descripcion": m.get("descripcion", ""),
                "activa": m.get("activa", True),
                "creada": m.get("creada"),
            })
        return jsonify({"listas": out})

    @app.route("/api/listas-precios", methods=["POST"])
    def api_listas_precios_crear():
        """Crea una nueva lista de precios (xlsx vacío con headers + entrada en catálogo)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from datetime import datetime as _dt
        from pathlib import Path
        import json as _json

        body = request.get_json(silent=True) or {}
        lista_id = (body.get("id") or "").strip()
        nombre = (body.get("nombre") or "").strip()
        descripcion = (body.get("descripcion") or "").strip()
        copiar_de = (body.get("copiar_de") or "").strip() or None  # opcional: clonar otra lista

        if not lista_id or not nombre:
            return jsonify({"ok": False, "error": "id y nombre son requeridos"}), 400
        # Sanitize ID
        import re
        if not re.match(r"^[A-Za-z0-9_-]+$", lista_id):
            return jsonify({"ok": False, "error": "id solo puede tener letras, números, _ o -"}), 400

        agentes_data = _cargar_agentes_data()
        archivos = agentes_data.get("listas_precios_archivos") or {}
        if lista_id in archivos:
            return jsonify({"ok": False, "error": f"ya existe una lista con id '{lista_id}'"}), 400

        rel_path = f"data/Lista_Precios_{lista_id}.xlsx"
        abs_path = Path(config.BASE_DIR) / rel_path
        abs_path.parent.mkdir(parents=True, exist_ok=True)

        # Crear xlsx (vacío o clonado)
        if copiar_de:
            src = _resolver_lista_path(copiar_de)
            if src and src.exists():
                import shutil
                shutil.copy2(src, abs_path)
            else:
                copiar_de = None
        if not copiar_de:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lista de Precios"
            headers = ["#", "Producto", "Unidad", "Precio Unitario"]
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E79")
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 16
            wb.save(abs_path)

        # Actualizar agentes.json
        archivos[lista_id] = rel_path
        meta = agentes_data.get("listas_precios_meta") or {}
        meta[lista_id] = {
            "nombre": nombre,
            "descripcion": descripcion,
            "activa": True,
            "creada": _dt.now().isoformat(timespec="seconds"),
        }
        agentes_data["listas_precios_archivos"] = archivos
        agentes_data["listas_precios_meta"] = meta
        _agentes_path().write_text(_json.dumps(agentes_data, ensure_ascii=False, indent=2),
                                     encoding="utf-8")
        log_event("system", f"💲 Nueva lista de precios creada: {lista_id}",
                  {"nombre": nombre, "copiada_de": copiar_de})
        return jsonify({"ok": True, "id": lista_id, "archivo": rel_path,
                        "items": (1 if copiar_de else 0)})

    @app.route("/api/listas-precios/<lista_id>/meta", methods=["POST"])
    def api_listas_precios_meta(lista_id):
        """Actualiza meta (nombre, descripción, activa) de una lista. Soft-delete via activa=false."""
        import json as _json
        body = request.get_json(silent=True) or {}
        agentes_data = _cargar_agentes_data()
        archivos = agentes_data.get("listas_precios_archivos") or {}
        if lista_id not in archivos:
            return jsonify({"ok": False, "error": "lista no existe"}), 404
        meta_all = agentes_data.get("listas_precios_meta") or {}
        m = meta_all.get(lista_id, {})
        for k in ("nombre", "descripcion", "activa"):
            if k in body:
                m[k] = body[k]
        meta_all[lista_id] = m
        agentes_data["listas_precios_meta"] = meta_all
        _agentes_path().write_text(_json.dumps(agentes_data, ensure_ascii=False, indent=2),
                                     encoding="utf-8")
        log_event("system", f"💲 Meta de lista '{lista_id}' actualizada", m)
        return jsonify({"ok": True, "meta": m})

    # ─── Lista de precios (editable, con multi-lista) ──────────────────────
    @app.route("/api/lista-precios", methods=["GET"])
    def api_lista_precios_get():
        """Lee los renglones de la lista pedida (?id=ID o default)."""
        import openpyxl
        lista_id = (request.args.get("id") or "").strip() or None
        path = _resolver_lista_path(lista_id)
        if not path or not path.exists():
            return jsonify({"error": "Lista no encontrada", "id": lista_id, "path": str(path)}), 404
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Lista de Precios"] if "Lista de Precios" in wb.sheetnames else wb.active
        headers = [c.value for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({
                "n": row[0] if len(row) > 0 else None,
                "producto": row[1] if len(row) > 1 else "",
                "unidad": row[2] if len(row) > 2 else "",
                "precio": float(row[3]) if len(row) > 3 and row[3] is not None else 0.0,
            })
        wb.close()
        return jsonify({
            "id": lista_id or "default",
            "path": str(path),
            "headers": headers,
            "items": rows,
            "total_items": len(rows),
        })

    @app.route("/api/lista-precios", methods=["POST"])
    def api_lista_precios_save():
        """Guarda la lista de precios completa desde el dashboard.

        Body JSON: {"items": [{"n": 1, "producto": "...", "unidad": "...", "precio": 25.5}, ...]}
        Crea backup antes de sobrescribir, recarga la cache de pricing.
        """
        import openpyxl
        from openpyxl.styles import Font
        from pathlib import Path
        from datetime import datetime as _dt
        import shutil

        body = request.get_json(silent=True) or {}
        items = body.get("items") or []
        lista_id = (body.get("id") or "").strip() or None
        if not items:
            return jsonify({"ok": False, "error": "No hay items para guardar"}), 400

        path = _resolver_lista_path(lista_id)
        if not path or not path.exists():
            return jsonify({"ok": False, "error": f"Lista no existe: {lista_id or 'default'}"}), 404

        # Backup automático
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        backup_path = path.with_suffix(f".backup_{ts}.xlsx")
        shutil.copy2(path, backup_path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Lista de Precios"] if "Lista de Precios" in wb.sheetnames else wb.active

        # Limpiar filas existentes (preservando header)
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            for c in range(1, 5):
                ws.cell(row=r, column=c).value = None

        # Escribir nuevas filas
        for i, it in enumerate(items, 1):
            ws.cell(row=i + 1, column=1, value=int(it.get("n") or i))
            ws.cell(row=i + 1, column=2, value=str(it.get("producto") or "").strip())
            ws.cell(row=i + 1, column=3, value=str(it.get("unidad") or "").strip())
            try:
                precio = float(it.get("precio") or 0)
            except (TypeError, ValueError):
                precio = 0.0
            ws.cell(row=i + 1, column=4, value=precio)
        wb.save(path)
        wb.close()

        # Recargar cache si es la lista default
        from .pricing import cargar_lista_precios
        n_loaded = 0
        if not lista_id or str(path) == str(Path(config.LISTA_PRECIOS_PATH)):
            cargar_lista_precios.cache_clear()
            n_loaded = len(cargar_lista_precios())

        log_event("system", f"💲 Lista '{lista_id or 'default'}' actualizada ({len(items)} items)",
                  {"id": lista_id, "backup": backup_path.name, "items": len(items),
                   "cargados_en_memoria": n_loaded})

        return jsonify({
            "ok": True,
            "id": lista_id,
            "items_guardados": len(items),
            "items_cargados": n_loaded,
            "backup": backup_path.name,
        })

    # ─── Agentes (multi-agente) ────────────────────────────────────────────
    def _agentes_path():
        from pathlib import Path
        return Path(config.BASE_DIR) / "storage" / "agentes.json"

    def _cargar_agentes_data() -> dict:
        import json as _json
        p = _agentes_path()
        if not p.exists():
            return {"agentes": [], "agente_default": None, "listas_precios_archivos": {}}
        try:
            return _json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {"agentes": [], "agente_default": None, "listas_precios_archivos": {}}

    def _resolver_agente(agent_id: str | None) -> dict | None:
        """Devuelve el agente solicitado, o el default si no se especifica.
        Si tampoco hay default, devuelve None (operación sigue con valores legacy).
        """
        data = _cargar_agentes_data()
        agentes = [a for a in data.get("agentes", []) if a.get("activo") is not False]
        if agent_id:
            for a in agentes:
                if a.get("id") == agent_id:
                    return a
        default_id = data.get("agente_default")
        if default_id:
            for a in agentes:
                if a.get("id") == default_id:
                    return a
        return agentes[0] if agentes else None

    @app.route("/api/agentes", methods=["GET"])
    def api_agentes_get():
        import json as _json
        p = _agentes_path()
        if not p.exists():
            return jsonify({"agentes": [], "agente_default": None})
        try:
            data = _json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            return jsonify({"error": f"agentes.json corrupto: {e}"}), 500
        return jsonify({
            "agentes": data.get("agentes", []),
            "agente_default": data.get("agente_default"),
            "listas_precios_archivos": data.get("listas_precios_archivos", {}),
        })

    @app.route("/api/agentes", methods=["POST"])
    def api_agentes_save():
        """Guarda el catálogo de agentes."""
        import json as _json
        body = request.get_json(silent=True) or {}
        agentes = body.get("agentes")
        if agentes is None:
            return jsonify({"ok": False, "error": "Falta 'agentes' en body"}), 400

        p = _agentes_path()
        existing = {}
        if p.exists():
            try:
                existing = _json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                pass

        new_data = {
            "_documentacion": existing.get("_documentacion",
                "Agentes especializados. Editable desde el dashboard."),
            "agente_default": body.get("agente_default", existing.get("agente_default")),
            "agentes": agentes,
            "listas_precios_archivos": body.get("listas_precios_archivos",
                existing.get("listas_precios_archivos", {})),
        }
        p.write_text(_json.dumps(new_data, ensure_ascii=False, indent=2), encoding="utf-8")
        log_event("system", f"🤖 Agentes actualizados ({len(agentes)} agentes)",
                  {"agentes": len(agentes), "default": new_data["agente_default"]})
        return jsonify({"ok": True, "agentes_guardados": len(agentes)})

    # ─── Clientes (editables) ──────────────────────────────────────────────
    def _clientes_path():
        from pathlib import Path
        return Path(config.BASE_DIR) / "storage" / "clientes.json"

    @app.route("/api/clientes", methods=["GET"])
    def api_clientes_get():
        import json as _json
        p = _clientes_path()
        if not p.exists():
            return jsonify({"clientes": [], "lista_precios_disponibles": []})
        try:
            data = _json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            return jsonify({"error": f"clientes.json corrupto: {e}"}), 500
        return jsonify({
            "clientes": data.get("clientes", []),
            "lista_precios_disponibles": data.get("lista_precios_disponibles", []),
        })

    @app.route("/api/clientes", methods=["POST"])
    def api_clientes_save():
        """Guarda el catálogo completo de clientes y listas disponibles."""
        import json as _json
        body = request.get_json(silent=True) or {}
        clientes = body.get("clientes")
        listas = body.get("lista_precios_disponibles")
        if clientes is None:
            return jsonify({"ok": False, "error": "Falta 'clientes' en body"}), 400

        p = _clientes_path()
        # Preservar campos meta (como _documentacion) si ya existían
        existing = {}
        if p.exists():
            try:
                existing = _json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                pass

        new_data = {
            "_documentacion": existing.get("_documentacion",
                "Catálogo editable de clientes. Editable desde el dashboard."),
            "lista_precios_disponibles": listas if listas is not None else
                existing.get("lista_precios_disponibles", []),
            "clientes": clientes,
        }
        p.write_text(_json.dumps(new_data, ensure_ascii=False, indent=2), encoding="utf-8")
        log_event("system", f"👥 Clientes actualizados desde dashboard ({len(clientes)} clientes)",
                  {"clientes": len(clientes)})
        return jsonify({"ok": True, "clientes_guardados": len(clientes)})

    # ─── Ventas (agregaciones para dashboard) ──────────────────────────────
    @app.route("/api/ventas")
    def api_ventas():
        """Agregaciones de ventas filtradas por rango de fechas.

        Params:
          desde=YYYY-MM-DD (default: primer día del mes actual)
          hasta=YYYY-MM-DD (default: hoy)
          agent_id=opcional
        """
        import json as _json
        from datetime import date as _date, datetime as _dt
        from pathlib import Path

        # Default: mes actual
        hoy = _date.today()
        primer_dia_mes = hoy.replace(day=1)
        desde_str = (request.args.get("desde") or "").strip() or primer_dia_mes.isoformat()
        hasta_str = (request.args.get("hasta") or "").strip() or hoy.isoformat()
        try:
            desde = _date.fromisoformat(desde_str)
            hasta = _date.fromisoformat(hasta_str)
        except ValueError:
            return jsonify({"error": "fechas inválidas (formato YYYY-MM-DD)"}), 400

        agent_id = (request.args.get("agent_id") or "").strip() or None

        pedidos_dir = Path(config.BASE_DIR) / "storage" / "pedidos_dia"
        extras_dir = Path(config.BASE_DIR) / "storage" / "extras_dia"

        # Map agent_id → cliente_id (si está pedido)
        cliente_filter = None
        if agent_id:
            data_ag = _cargar_agentes_data()
            for a in data_ag.get("agentes", []):
                if a.get("id") == agent_id:
                    cliente_filter = a.get("cliente_id")
                    break

        dias = []                       # serie temporal (para gráfica)
        por_hospital: dict = {}        # totales por destino
        remisiones: list = []           # detalle plano
        total_dia_general = 0.0
        n_remisiones = 0
        total_regular = 0.0
        total_extras = 0.0

        # Iterar fechas en rango
        from datetime import timedelta as _td
        d = desde
        while d <= hasta:
            iso = d.isoformat()
            dia_total = 0.0
            dia_reg = 0.0
            dia_ext = 0.0

            sp = pedidos_dir / f"{iso}.json"
            if sp.exists():
                try:
                    state = _json.loads(sp.read_text(encoding="utf-8"))
                    for hospital, info in state.get("hospitales", {}).items():
                        # Si filtramos por cliente, solo SUREÑA es comedores; EHMO incluye
                        # todo lo regular. Para v1 si cliente_filter es SURENA saltamos
                        # los hospitales (pertenecen a EHMO).
                        if cliente_filter == "SURENA" and not (
                            "comedor" in hospital.lower() or "almacén" in hospital.lower()
                        ):
                            continue
                        total = info.get("total", 0)
                        if total <= 0:
                            continue
                        dia_total += total; dia_reg += total
                        n_remisiones += 1
                        por_hospital[hospital] = por_hospital.get(hospital, 0.0) + total
                        remisiones.append({
                            "fecha_iso": iso,
                            "destino": hospital,
                            "tipo": "regular",
                            "estado": info.get("estado", "vigente"),
                            "folio": info.get("folio_remision"),
                            "total": round(total, 2),
                        })
                except Exception:
                    pass

            ep = extras_dir / f"{iso}.json"
            if ep.exists():
                try:
                    ex = _json.loads(ep.read_text(encoding="utf-8"))
                    folios_dest = ex.get("folios_por_destino") or {}
                    estados_dest = ex.get("estados_por_destino") or {}
                    grupos: dict[str, list] = {}
                    for e in ex.get("extras", []):
                        if e.get("cantidad", 0) <= 0:
                            continue
                        grupos.setdefault(e.get("hospital", "ALMACÉN EHMO"), []).append(e)
                    for destino, items in grupos.items():
                        total = sum(it.get("importe", 0) for it in items)
                        dia_total += total; dia_ext += total
                        n_remisiones += 1
                        nombre = f"{destino} (EXTRA)"
                        por_hospital[nombre] = por_hospital.get(nombre, 0.0) + total
                        remisiones.append({
                            "fecha_iso": iso,
                            "destino": nombre,
                            "tipo": "extra",
                            "estado": estados_dest.get(destino, "vigente"),
                            "folio": folios_dest.get(destino),
                            "total": round(total, 2),
                        })
                except Exception:
                    pass

            dias.append({
                "fecha": iso,
                "total": round(dia_total, 2),
                "regular": round(dia_reg, 2),
                "extras": round(dia_ext, 2),
            })
            total_dia_general += dia_total
            total_regular += dia_reg
            total_extras += dia_ext
            d = d + _td(days=1)

        # Top hospitales (orden desc por monto)
        top_hospitales = sorted(
            [{"destino": h, "total": round(v, 2)} for h, v in por_hospital.items()],
            key=lambda x: -x["total"],
        )

        return jsonify({
            "desde": desde_str,
            "hasta": hasta_str,
            "agent_id": agent_id,
            "kpi": {
                "total": round(total_dia_general, 2),
                "regular": round(total_regular, 2),
                "extras": round(total_extras, 2),
                "n_remisiones": n_remisiones,
                "n_destinos": len(por_hospital),
                "ticket_promedio": round(total_dia_general / n_remisiones, 2) if n_remisiones else 0,
                "n_dias_con_ventas": sum(1 for d in dias if d["total"] > 0),
            },
            "serie_diaria": dias,
            "top_hospitales": top_hospitales,
            "remisiones_count": len(remisiones),
        })

    @app.route("/files/processed/<path:filename>")
    def serve_processed(filename):
        """Sirve archivos generados desde storage/processed/ para descargar/ver."""
        from flask import send_from_directory, abort
        directory = Path(config.PROCESSED_DIR).resolve()
        # Validación: el archivo debe existir Y estar dentro del directorio (no ../)
        target = (directory / filename).resolve()
        try:
            target.relative_to(directory)
        except ValueError:
            return abort(403)
        if not target.exists() or not target.is_file():
            return abort(404)
        return send_from_directory(str(directory), filename)

    return app


app = create_app()


if __name__ == "__main__":
    config.validate_config()
    log.info(f"🚀 Frutas Kelly WhatsApp Agent — entorno: {config.ENVIRONMENT}")
    log.info(f"   Webhook URL local: http://localhost:{config.PORT}/webhook")
    log.info(f"   Verify token: {config.WHATSAPP_VERIFY_TOKEN}")
    app.run(host="0.0.0.0", port=config.PORT, debug=(config.ENVIRONMENT == "development"))
