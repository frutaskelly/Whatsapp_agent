"""Genera la 'Relación de documentos' por semana.

Replica el formato del Excel del cliente:
  FACTURACION DE HOSPITALES CHIAPAS EHMO SEM<n>.xlsx

Estructura del Excel resultante:
  - Hoja "S": resumen general de la semana
  - Una hoja por día (LUNES, MARTES, MIERCOLES, ...)

Cada hoja tiene:
  - Fila 1: vacía
  - Fila 2: título "REMISION Y FACTURACION DE HOSPITALES CHIAPAS PROYECTO EHMO"
  - Fila 4: headers (#, SEMANA, FECHA, HOSPITAL, # REMISION/PEDIDO, TOTAL, FECHA ELAB, # FACTURA, TOTAL FAC, FECHA ELAB FAC, OBSERVACIONES)
  - Filas 5+: una por hospital del catálogo, con datos del estado del día

Datos vienen de:
  - storage/pedidos_dia/<fecha-iso>.json (estado de hospitales con folio + total)
  - storage/extras_dia/<fecha-iso>.json (extras, mostrados aparte si aplica)
"""
import logging
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table as RlTable, TableStyle as RlTableStyle,
    Paragraph, Spacer,
)

from . import config
from .event_log import log_event
from .estado_pedido import cargar_estado
from .pedido_processor import HOSPITALES_CONOCIDOS_SI

log = logging.getLogger(__name__)

# ─── Estilos ──────────────────────────────────────────────────────────────────
AZUL_OSC = "1F4E79"
AZUL_CLAR = "D6E4F0"
AMARILLO = "FFF2CC"
GRIS = "BBBBBB"

DIAS_NOMBRE = {0: "LUNES", 1: "MARTES", 2: "MIERCOLES", 3: "JUEVES",
               4: "VIERNES", 5: "SABADO", 6: "DOMINGO"}
MESES_NOMBRE = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}

# Offset entre semana del cliente EHMO y semana ISO (de nota_remision)
SEMANA_OFFSET_EHMO = -1


def _semana_a_iso(semana_ehmo: int) -> int:
    return semana_ehmo - SEMANA_OFFSET_EHMO  # invertir el offset


def _lunes_de_semana(semana_ehmo: int, year: int) -> datetime:
    """Devuelve el lunes (de la semana ISO equivalente) para una semana EHMO."""
    iso_week = _semana_a_iso(semana_ehmo)
    return datetime.strptime(f"{year}-W{iso_week:02d}-1", "%G-W%V-%u")


def _border():
    s = Side(style="thin", color=GRIS)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_cell(ws, row, col, value, *, bold=False, bg=None, color="000000",
              size=10, align="left", number_format=None, border_=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=size, color=color)
    if border_:
        c.border = _border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    return c


def _setup_hoja_dia(ws, semana_ehmo: int, fecha_legible: str,
                    hospitales_canonicos: list[str]):
    """Configura el header de una hoja de día (sin datos, solo estructura)."""
    # Título general
    ws.merge_cells("C2:K2")
    _set_cell(ws, 2, 3, "REMISION Y FACTURACION DE HOSPITALES CHIAPAS PROYECTO EHMO",
              bold=True, bg=AZUL_OSC, color="FFFFFF", size=12, align="center")

    # Headers
    headers = ["#", "SEMANA", "FECHA", "HOSPITAL",
               "# REMISIÓN / PEDIDO", "TOTAL DEL PEDIDO", "FECHA ELABORACIÓN",
               "# FACTURA", "TOTAL FACTURA", "FECHA ELABORACIÓN", "OBSERVACIONES"]
    for col_idx, h in enumerate(headers, start=2):
        _set_cell(ws, 4, col_idx, h, bold=True, bg=AZUL_OSC, color="FFFFFF",
                  align="center", size=10)

    # Una fila por hospital del catálogo
    semana_label = f"SEMANA {semana_ehmo}"
    for i, hospital in enumerate(hospitales_canonicos, 1):
        row = 4 + i
        _set_cell(ws, row, 2, i, align="center")
        _set_cell(ws, row, 3, semana_label, align="center")
        _set_cell(ws, row, 4, fecha_legible, align="center")
        _set_cell(ws, row, 5, hospital)
        # 6-12 quedan vacíos (folio, total, fechas, factura, obs)
        for c in range(6, 13):
            _set_cell(ws, row, c, None)

    # Anchos de columna
    widths = [3, 5, 12, 15, 60, 18, 16, 18, 14, 16, 18, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _llenar_datos_dia(ws, fecha_iso: str, hospitales_canonicos: list[str]):
    """Llena los datos del día desde el estado pedidos_dia/<fecha-iso>.json."""
    estado = cargar_estado(fecha_iso)
    if not estado:
        return 0
    hospitales_estado = estado.get("hospitales", {})
    rows_llenas = 0
    for i, hospital in enumerate(hospitales_canonicos, 1):
        info = None
        # Solo match exacto por nombre canónico (los nombres del estado YA son
        # canónicos porque vienen de la BD original; el fuzzy aquí causa
        # falsos positivos como "Berriozabal" matcheando "Chiapa de Corzo"
        # por palabras genéricas en común — "Hospital", "Básico", etc.).
        if hospital in hospitales_estado:
            info = hospitales_estado[hospital]
        if not info or info.get("total", 0) <= 0:
            continue
        row = 4 + i
        folio = info.get("folio_remision") or ""
        if folio:
            try:
                folio = int(folio)  # mostrar como número (313 en vez de "0000000313")
            except ValueError:
                pass
        _set_cell(ws, row, 6, folio, align="center")
        _set_cell(ws, row, 7, info.get("total", 0), align="right",
                  number_format='"$"#,##0.00')
        _set_cell(ws, row, 8, fecha_iso, align="center")
        rows_llenas += 1
    return rows_llenas


def _setup_hoja_resumen(ws, semana_ehmo: int, fecha_inicio: datetime,
                        hospitales_canonicos: list[str]):
    """Hoja 'S' — resumen de la semana, totales por hospital."""
    ws.title = "S"

    ws.merge_cells("C2:K2")
    _set_cell(ws, 2, 3, "REMISION Y FACTURACION DE HOSPITALES CHIAPAS PROYECTO EHMO",
              bold=True, bg=AZUL_OSC, color="FFFFFF", size=12, align="center")

    headers = ["#", "SEMANA", "RANGO", "HOSPITAL",
               "TOTAL SEMANA", "# REMISIONES", "OBSERVACIONES"]
    for col_idx, h in enumerate(headers, start=2):
        _set_cell(ws, 4, col_idx, h, bold=True, bg=AZUL_OSC, color="FFFFFF",
                  align="center", size=10)

    rango = f"{fecha_inicio.day}–{(fecha_inicio + timedelta(days=6)).day} de " \
            f"{MESES_NOMBRE[fecha_inicio.month]}"
    semana_label = f"SEMANA {semana_ehmo}"

    totales_por_hospital = {h: 0.0 for h in hospitales_canonicos}
    remisiones_por_hospital = {h: 0 for h in hospitales_canonicos}

    for i in range(7):
        f = fecha_inicio + timedelta(days=i)
        fecha_iso = f.strftime("%Y-%m-%d")
        estado = cargar_estado(fecha_iso)
        if not estado:
            continue
        for hospital_canonico in hospitales_canonicos:
            # Solo match exacto (evita falsos positivos por palabras genéricas)
            info = estado["hospitales"].get(hospital_canonico)
            if info and info.get("total", 0) > 0:
                totales_por_hospital[hospital_canonico] += info["total"]
                remisiones_por_hospital[hospital_canonico] += 1

    total_semana = sum(totales_por_hospital.values())
    total_remisiones = sum(remisiones_por_hospital.values())

    for i, hospital in enumerate(hospitales_canonicos, 1):
        row = 4 + i
        _set_cell(ws, row, 2, i, align="center")
        _set_cell(ws, row, 3, semana_label, align="center")
        _set_cell(ws, row, 4, rango, align="center")
        _set_cell(ws, row, 5, hospital)
        total = totales_por_hospital[hospital]
        rems = remisiones_por_hospital[hospital]
        _set_cell(ws, row, 6, total if total else None, align="right",
                  number_format='"$"#,##0.00')
        _set_cell(ws, row, 7, rems if rems else None, align="center")
        _set_cell(ws, row, 8, None)

    # Total general
    last_row = 4 + len(hospitales_canonicos) + 1
    _set_cell(ws, last_row, 5, "TOTAL SEMANA", bold=True, bg=AMARILLO, align="right")
    _set_cell(ws, last_row, 6, total_semana, bold=True, bg=AMARILLO,
              align="right", number_format='"$"#,##0.00')
    _set_cell(ws, last_row, 7, total_remisiones, bold=True, bg=AMARILLO,
              align="center")

    widths = [3, 5, 12, 15, 60, 18, 14, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _cargar_extras_para_relacion(fecha_iso: str) -> list[dict]:
    """Devuelve los destinos de extras del día con sus totales y folios.

    Útil para incluir en la relación de documentos (no solo hospitales del
    pedido normal, sino también el ALMACÉN EHMO u otros destinos de extras).
    """
    try:
        from .extras_pedido import cargar_extras
        state = cargar_extras(fecha_iso)
    except Exception:
        return []
    if not state:
        return []
    extras = state.get("extras", [])
    folios = state.get("folios_por_destino") or {}
    # Agrupar por destino
    por_destino = {}
    for e in extras:
        if e.get("cantidad", 0) <= 0:
            continue
        h = e.get("hospital", "ALMACÉN EHMO")
        por_destino.setdefault(h, {"total": 0, "items": 0})
        por_destino[h]["total"] += e.get("importe", 0)
        por_destino[h]["items"] += 1
    return [{"destino": d, "total": v["total"], "items": v["items"],
             "folio": folios.get(d)} for d, v in sorted(por_destino.items())]


def generar_relacion_dia(fecha_iso: str,
                          fecha_legible: str | None = None,
                          output_path: Path | None = None) -> dict:
    """Genera Excel de Relación de Documentos para UN DÍA específico.

    Incluye los hospitales que tuvieron pedido FyV ese día (con folio y total)
    + los destinos de extras del día (ej. ALMACÉN EHMO).
    """
    estado = cargar_estado(fecha_iso)
    if not estado:
        return {"error": f"No hay estado del día {fecha_iso}"}

    fecha_legible = fecha_legible or estado.get("fecha_legible", fecha_iso)
    fecha_dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
    semana_ehmo = max(1, fecha_dt.isocalendar()[1] + SEMANA_OFFSET_EHMO)

    # Solo hospitales con total > 0
    hospitales_activos = sorted(
        [(h, info) for h, info in estado["hospitales"].items()
         if info.get("total", 0) > 0],
        key=lambda x: x[0],
    )
    extras_destinos = _cargar_extras_para_relacion(fecha_iso)

    if not hospitales_activos and not extras_destinos:
        return {"error": f"Sin hospitales con pedido FyV el {fecha_iso}"}

    wb = openpyxl.Workbook()
    ws = wb.active
    nombre_dia = DIAS_NOMBRE[fecha_dt.weekday()]
    ws.title = f"SEM{semana_ehmo} {nombre_dia} {fecha_dt.day}"[:31]

    # Header
    ws.merge_cells("C2:K2")
    _set_cell(ws, 2, 3,
              "REMISION Y FACTURACION DE HOSPITALES CHIAPAS PROYECTO EHMO",
              bold=True, bg=AZUL_OSC, color="FFFFFF", size=12, align="center")

    headers = ["#", "SEMANA", "FECHA", "HOSPITAL",
               "# REMISIÓN / PEDIDO", "TOTAL DEL PEDIDO", "FECHA ELABORACIÓN",
               "ESTATUS",
               "# FACTURA", "TOTAL FACTURA", "FECHA ELABORACIÓN", "OBSERVACIONES"]
    for col_idx, h in enumerate(headers, start=2):
        _set_cell(ws, 4, col_idx, h, bold=True, bg=AZUL_OSC, color="FFFFFF",
                  align="center", size=10)

    semana_label = f"SEMANA {semana_ehmo}"
    fecha_dia_legible = f"{fecha_dt.day} DE {MESES_NOMBRE[fecha_dt.month]}"

    # Mapeo de estado → texto y color de fondo
    ESTADO_DISPLAY = {
        "vigente": ("🆕 Vigente", None),
        "modificado": ("🔄 Modificado", "FFF2CC"),
        "aceptado": ("✅ Aceptado", "C6EFCE"),
        "cancelado": ("❌ Cancelado", "FFC7CE"),
    }

    total_general = 0.0
    row_idx = 0
    for i, (hospital, info) in enumerate(hospitales_activos, 1):
        row = 4 + i
        row_idx = i
        _set_cell(ws, row, 2, i, align="center")
        _set_cell(ws, row, 3, semana_label, align="center")
        _set_cell(ws, row, 4, fecha_dia_legible, align="center")
        _set_cell(ws, row, 5, hospital)
        folio = info.get("folio_remision") or ""
        if folio:
            try:
                folio = int(folio)
            except (ValueError, TypeError):
                pass
        _set_cell(ws, row, 6, folio, align="center")
        _set_cell(ws, row, 7, info.get("total", 0), align="right",
                  number_format='"$"#,##0.00')
        _set_cell(ws, row, 8, fecha_iso, align="center")
        # Columna ESTATUS (col 9)
        estado = info.get("estado", "vigente")
        est_text, est_bg = ESTADO_DISPLAY.get(estado, (estado, None))
        _set_cell(ws, row, 9, est_text, align="center", bg=est_bg, bold=True)
        total_general += info.get("total", 0)
        # Las 4 columnas de factura (col 10-13) quedan vacías para llenar después

    # Filas adicionales para destinos de extras (ALMACÉN EHMO, etc.)
    # Cargar estados de extras
    estados_extras = {}
    try:
        from .extras_pedido import cargar_extras
        ex_state = cargar_extras(fecha_iso)
        if ex_state:
            estados_extras = ex_state.get("estados_por_destino") or {}
    except Exception:
        pass
    for ext in extras_destinos:
        row_idx += 1
        row = 4 + row_idx
        _set_cell(ws, row, 2, row_idx, align="center")
        _set_cell(ws, row, 3, semana_label, align="center")
        _set_cell(ws, row, 4, fecha_dia_legible, align="center")
        _set_cell(ws, row, 5, f"{ext['destino']} (EXTRA)", bold=True)
        folio = ext.get("folio") or ""
        if folio:
            try:
                folio = int(folio)
            except (ValueError, TypeError):
                pass
        _set_cell(ws, row, 6, folio, align="center")
        _set_cell(ws, row, 7, ext["total"], align="right",
                  number_format='"$"#,##0.00')
        _set_cell(ws, row, 8, fecha_iso, align="center")
        # Estatus para extras
        estado = estados_extras.get(ext["destino"], "vigente")
        est_text, est_bg = ESTADO_DISPLAY.get(estado, (estado, None))
        _set_cell(ws, row, 9, est_text, align="center", bg=est_bg, bold=True)
        _set_cell(ws, row, 13, f"Cubrir desabasto ({ext['items']} productos)")
        total_general += ext["total"]

    # Fila de total general (ahora col 7 = TOTAL, ajustar)
    last = 4 + row_idx + 1
    _set_cell(ws, last, 6, "TOTAL DEL DÍA", bold=True, bg=AMARILLO, align="right")
    _set_cell(ws, last, 7, total_general, bold=True, bg=AMARILLO, align="right",
              number_format='"$"#,##0.00')

    # Anchos
    widths = [3, 5, 12, 15, 60, 18, 16, 16, 16, 14, 18, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if not output_path:
        output_path = (config.PROCESSED_DIR /
                       f"Relación Documentos {fecha_legible} ({fecha_iso}).xlsx")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    log.info(f"Relación día generada: {output_path}")
    log_event("processor",
              f"📋 Relación documentos del día {fecha_iso} ({len(hospitales_activos)} hospitales, ${total_general:,.2f})",
              {"output": output_path.name,
               "hospitales": len(hospitales_activos),
               "total": total_general})

    return {
        "output_path": output_path,
        "fecha": fecha_iso,
        "fecha_legible": fecha_legible,
        "hospitales_count": len(hospitales_activos),
        "total_general": total_general,
    }


def generar_relacion_dia_pdf(fecha_iso: str,
                              fecha_legible: str | None = None,
                              output_path: Path | None = None) -> dict:
    """Genera PDF (landscape) de Relación de Documentos del día — para imprimir.

    Mismo contenido que el Excel pero en una hoja horizontal lista para
    imprimir. Solo incluye hospitales con pedido FyV ese día.
    """
    estado = cargar_estado(fecha_iso)
    if not estado:
        return {"error": f"No hay estado del día {fecha_iso}"}

    fecha_legible = fecha_legible or estado.get("fecha_legible", fecha_iso)
    fecha_dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
    semana_ehmo = max(1, fecha_dt.isocalendar()[1] + SEMANA_OFFSET_EHMO)

    hospitales_activos = sorted(
        [(h, info) for h, info in estado["hospitales"].items()
         if info.get("total", 0) > 0],
        key=lambda x: x[0],
    )
    extras_destinos = _cargar_extras_para_relacion(fecha_iso)
    if not hospitales_activos and not extras_destinos:
        return {"error": f"Sin hospitales con pedido FyV el {fecha_iso}"}

    if not output_path:
        output_path = (config.PROCESSED_DIR /
                       f"Relación Documentos {fecha_legible} ({fecha_iso}).pdf")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ─── Estilos PDF ─────────────────────────────────────────────────────────
    rl_AZUL_OSC = rl_colors.HexColor(f"#{AZUL_OSC}")
    rl_AZUL_CLAR = rl_colors.HexColor(f"#{AZUL_CLAR}")
    rl_AMARILLO = rl_colors.HexColor(f"#{AMARILLO}")
    rl_GRIS = rl_colors.HexColor(f"#{GRIS}")

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(letter),
        topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        leftMargin=0.8 * cm, rightMargin=0.8 * cm,
        title=f"Relación Documentos {fecha_legible}",
    )

    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle(
        "tit", parent=styles["Heading1"],
        fontSize=13, leading=16, alignment=TA_CENTER, textColor=rl_AZUL_OSC,
        spaceAfter=2, fontName="Helvetica-Bold",
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=11, leading=13, alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#444"), spaceAfter=8,
    )

    elements = [
        Paragraph("REMISIÓN Y FACTURACIÓN DE HOSPITALES CHIAPAS — PROYECTO EHMO",
                  style_titulo),
        Paragraph(
            f"SEMANA {semana_ehmo} &nbsp;·&nbsp; "
            f"{DIAS_NOMBRE[fecha_dt.weekday()]} "
            f"{fecha_dt.day} DE {MESES_NOMBRE[fecha_dt.month]} {fecha_dt.year}",
            style_sub,
        ),
    ]

    # ─── Tabla ───────────────────────────────────────────────────────────────
    headers = ["#", "SEMANA", "FECHA", "HOSPITAL",
               "# REMISIÓN /\nPEDIDO", "TOTAL DEL\nPEDIDO",
               "FECHA\nELAB.", "ESTATUS", "# FACTURA",
               "TOTAL\nFACTURA", "FECHA\nELAB. FAC.", "OBSERVACIONES"]
    data = [headers]

    semana_label = f"SEMANA {semana_ehmo}"
    fecha_dia_legible = f"{fecha_dt.day} DE {MESES_NOMBRE[fecha_dt.month]}"

    # Mapeo de estado → texto y color de fondo (PDF)
    ESTADO_DISPLAY_PDF = {
        "vigente": ("🆕 Vigente", None),
        "modificado": ("🔄 Modificado", rl_colors.HexColor("#FFF2CC")),
        "aceptado": ("✅ Aceptado", rl_colors.HexColor("#C6EFCE")),
        "cancelado": ("❌ Cancelado", rl_colors.HexColor("#FFC7CE")),
    }

    # Cargar estados de extras una sola vez
    estados_extras = {}
    try:
        from .extras_pedido import cargar_extras
        ex_state = cargar_extras(fecha_iso)
        if ex_state:
            estados_extras = ex_state.get("estados_por_destino") or {}
    except Exception:
        pass

    # Trackear el estado por fila para aplicar BACKGROUND en TableStyle
    estatus_row_styles = []  # [(row_idx, bg_color), ...]

    total_general = 0.0
    idx = 0
    for i, (hospital, info) in enumerate(hospitales_activos, 1):
        idx = i
        folio = info.get("folio_remision") or ""
        if folio:
            try:
                folio = str(int(folio))
            except (TypeError, ValueError):
                folio = str(folio)
        total = info.get("total", 0)
        total_general += total
        estado_h = info.get("estado", "vigente")
        est_text, est_bg = ESTADO_DISPLAY_PDF.get(estado_h, (estado_h, None))
        if est_bg is not None:
            estatus_row_styles.append((i, est_bg))
        data.append([
            str(i),
            semana_label,
            fecha_dia_legible,
            hospital,
            folio,
            f"${total:,.2f}",
            fecha_iso,
            est_text,
            "",   # # factura
            "",   # total factura
            "",   # fecha elab factura
            "",   # observaciones
        ])

    # Filas para extras (ALMACÉN EHMO, etc.)
    for ext in extras_destinos:
        idx += 1
        folio = ext.get("folio") or ""
        if folio:
            try:
                folio = str(int(folio))
            except (TypeError, ValueError):
                folio = str(folio)
        total_general += ext["total"]
        estado_e = estados_extras.get(ext["destino"], "vigente")
        est_text, est_bg = ESTADO_DISPLAY_PDF.get(estado_e, (estado_e, None))
        if est_bg is not None:
            estatus_row_styles.append((idx, est_bg))
        data.append([
            str(idx),
            semana_label,
            fecha_dia_legible,
            f"{ext['destino']} (EXTRA)",
            folio,
            f"${ext['total']:,.2f}",
            fecha_iso,
            est_text,
            "", "", "",
            f"Cubrir desabasto ({ext['items']} productos)",
        ])

    # Fila de total
    data.append([
        "", "", "", "TOTAL DEL DÍA",
        "", f"${total_general:,.2f}",
        "", "", "", "", "", "",
    ])

    # Anchos de columna optimizados para letter horizontal (~26.3 cm útil)
    col_widths = [0.7, 1.7, 1.7, 6.5, 1.9, 2.1, 1.6, 2.0, 1.6, 1.7, 1.7, 2.7]
    table = RlTable(
        data, colWidths=[w * cm for w in col_widths], repeatRows=1,
    )
    table_style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), rl_AZUL_OSC),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        # Body
        ("FONTSIZE", (0, 1), (-1, -2), 8),
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("ALIGN", (7, 1), (7, -1), "CENTER"),
        ("ALIGN", (9, 1), (9, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_GRIS),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
            [rl_colors.white, rl_colors.HexColor("#F8F9FB")]),
        # Total row
        ("BACKGROUND", (0, -1), (-1, -1), rl_AMARILLO),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 9),
        ("ALIGN", (5, -1), (5, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, rl_AZUL_OSC),
    ]
    # Aplicar background del estatus por fila (col 7 = índice de columna ESTATUS)
    for row_i, bg_color in estatus_row_styles:
        table_style_cmds.append(("BACKGROUND", (7, row_i), (7, row_i), bg_color))
        table_style_cmds.append(("FONTNAME", (7, row_i), (7, row_i), "Helvetica-Bold"))

    table.setStyle(RlTableStyle(table_style_cmds))
    elements.append(table)

    doc.build(elements)
    log.info(f"Relación día PDF generada: {output_path}")
    log_event("processor",
              f"📋 Relación de documentos PDF (horizontal) — {len(hospitales_activos)} hospitales",
              {"output": output_path.name, "hospitales": len(hospitales_activos)})

    return {
        "output_path": output_path,
        "fecha": fecha_iso,
        "hospitales_count": len(hospitales_activos),
        "total_general": total_general,
    }


def generar_relacion_semanal(semana_ehmo: int, year: int | None = None,
                              output_path: Path | None = None) -> dict:
    """Genera el Excel de Relación de Documentos para una semana completa."""
    year = year or datetime.now().year
    lunes = _lunes_de_semana(semana_ehmo, year)
    domingo = lunes + timedelta(days=6)

    hospitales_canonicos = list(HOSPITALES_CONOCIDOS_SI.keys())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja resumen
    ws_s = wb.create_sheet("S")
    _setup_hoja_resumen(ws_s, semana_ehmo, lunes, hospitales_canonicos)

    # Hoja por día
    dias_con_data = 0
    for i in range(7):
        f = lunes + timedelta(days=i)
        fecha_iso = f.strftime("%Y-%m-%d")
        nombre_dia = DIAS_NOMBRE[f.weekday()]
        sheet_name = f"SEM{semana_ehmo} {nombre_dia} {f.day}"[:31]
        ws = wb.create_sheet(sheet_name)
        fecha_legible = f"{f.day} DE {MESES_NOMBRE[f.month]}"
        _setup_hoja_dia(ws, semana_ehmo, fecha_legible, hospitales_canonicos)
        rows = _llenar_datos_dia(ws, fecha_iso, hospitales_canonicos)
        if rows:
            dias_con_data += 1

    if not output_path:
        output_path = (config.PROCESSED_DIR /
                       f"Relación de Documentos SEM{semana_ehmo} ({lunes.strftime('%d')}–"
                       f"{domingo.strftime('%d')} {MESES_NOMBRE[lunes.month]}).xlsx")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Relación generada: {output_path}")
    log_event("processor", f"📋 Relación de documentos SEM{semana_ehmo} generada",
              {"output": output_path.name, "dias_con_data": dias_con_data,
               "total_hospitales": len(hospitales_canonicos)})

    return {
        "output_path": output_path,
        "semana": semana_ehmo,
        "year": year,
        "dias_con_data": dias_con_data,
        "rango_inicio": lunes.strftime("%Y-%m-%d"),
        "rango_fin": domingo.strftime("%Y-%m-%d"),
    }
